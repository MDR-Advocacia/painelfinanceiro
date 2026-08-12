# Carrega as ABAS MENSAIS da planilha do DP (Jan..Dez) como competências.
#
#   python manage.py carregar_folhas_planilha --arquivo /tmp/folha.xlsx --ate 8 --fechar-ate 7
#
# O importador de cadastro (endpoint /dp/importar/) só lê TB_Colaboradores,
# TB_Cargos, CONFIG e Desligados. As abas de mês ficavam de fora — e são elas
# que têm o histórico: faltas, prêmios, acertos e o salário VIGENTE NAQUELE MÊS,
# que muitas vezes difere do cadastro de hoje.
#
# O QUE É ENTRADA E O QUE É CÁLCULO: da aba mensal só lemos o que o DP DIGITA
# (salário, VT, VA, faltas, saldo livre, acerto, premiações). Tudo o mais na
# planilha — desconto de faltas, INSS, provisões, custo total — é fórmula, e
# quem recalcula é o motor do painel. Por isso o comando compara, no fim, o
# TOTAL e o CUSTO que a planilha calculou com os que o motor calculou: é a
# prova de que o espelho está fiel.
#
# COLUNAS POR NOME, não por posição: as abas variam de largura entre os meses
# (Jan tem 35 colunas, Jul tem 36) e uma mudança de layout não pode fazer o
# comando ler salário na coluna do VT em silêncio.
import unicodedata
from datetime import date

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from financeiro.dp_folha import calcular_item, tabela_fiscal_para
from financeiro.models import (
    DpAuditLog, DpColaborador, DpCompetencia, DpFolhaItem, DpLancamento,
)

MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago",
         "Set", "Out", "Nov", "Dez"]

USUARIO = "carga histórica (planilha do DP)"


def _txt(v) -> str:
    return str(v).strip() if v is not None else ""


def _chave(s: str) -> str:
    """Normaliza cabeçalho pra casar sem depender de acento, caixa ou pontuação."""
    s = unicodedata.normalize("NFKD", _txt(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "".join(c for c in s.lower() if c.isalnum())


# nome do campo → pedaços que identificam a coluna no cabeçalho
COLUNAS = {
    "matricula":     ["matcodcolaborador", "matricula"],
    "salario":       ["salbrutor", "salariobrutor"],
    "vt":            ["vtr"],
    "va":            ["var"],
    "faltas_dias":   ["faltasemdias"],
    "faltas_horas":  ["faltasemhoras"],
    "saldo_livre":   ["saldolivrer"],
    "acerto":        ["acertocontabilr"],
    "premiacoes":    ["premiacoesextrar", "premiacoesextrasr"],
    # colunas CALCULADAS pela planilha — só para conferência.
    # ATENÇÃO ao nome: "Custo Total Mensal" da planilha NÃO é o custo total —
    # é só a soma das PROVISÕES (13º + férias + 1/3 + FGTS + multa + recesso).
    # Comparar com o nosso custo_total daria diferença enorme e falsa; o par
    # correto é com custo_provisoes.
    "total_planilha": ["totalr"],
    "provisoes_planilha": ["custototalmensal", "custototalmensalr"],
}


def _mapear(header) -> dict:
    """cabeçalho → {campo: índice}. Casamento exato primeiro, prefixo depois."""
    norm = [_chave(h) for h in header]
    mapa = {}
    for campo, alvos in COLUNAS.items():
        for alvo in alvos:
            if alvo in norm:
                mapa[campo] = norm.index(alvo)
                break
        if campo in mapa:
            continue
        for i, h in enumerate(norm):
            if h and any(h.startswith(a) for a in alvos):
                mapa[campo] = i
                break
    return mapa


def _num(v, padrao=0.0):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return padrao


class Command(BaseCommand):
    help = "Carrega as abas mensais da planilha do DP como competências da folha."

    def add_arguments(self, p):
        p.add_argument("--arquivo", required=True)
        p.add_argument("--ano", type=int, default=2026)
        p.add_argument("--de", type=int, default=1, help="primeiro mês (1-12)")
        p.add_argument("--ate", type=int, default=12, help="último mês (1-12)")
        p.add_argument("--fechar-ate", type=int, default=0,
                       help="fecha as competências até este mês (0 = não fecha)")
        p.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **o):
        from openpyxl import load_workbook

        dry = o["dry_run"]
        ano = o["ano"]
        wb = load_workbook(o["arquivo"], read_only=True, data_only=True)
        por_matricula = {c.matricula: c for c in
                         DpColaborador.objects.select_related("cargo", "centro_custo")
                         .prefetch_related("dependentes")}
        self.stdout.write(f"colaboradores no banco: {len(por_matricula)}")

        resumo = []
        for mes in range(o["de"], o["ate"] + 1):
            aba = MESES[mes - 1]
            if aba not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"  {aba}: aba inexistente — pulando"))
                continue
            ws = wb[aba]

            # linha 2 traz dias do mês e dias úteis; linha 3 é o cabeçalho
            dias_mes, dias_uteis = 30, 22
            header = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), start=1):
                vals = [_txt(c) for c in row]
                for j, v in enumerate(vals):
                    if _chave(v) == "diasdomes" and j + 1 < len(row):
                        dias_mes = int(_num(row[j + 1], 30) or 30)
                    if _chave(v).startswith("diasuteisdomes") and j + 1 < len(row):
                        dias_uteis = int(_num(row[j + 1], 22) or 22)
                if any(_chave(v).startswith("matcodcolaborador") for v in vals):
                    header = list(row)
                    linha_header = i
            if not header:
                self.stdout.write(self.style.ERROR(f"  {aba}: cabeçalho não encontrado — pulando"))
                continue
            mapa = _mapear(header)
            faltando = [c for c in ("matricula", "salario") if c not in mapa]
            if faltando:
                self.stdout.write(self.style.ERROR(
                    f"  {aba}: colunas essenciais ausentes {faltando} — pulando"))
                continue

            linhas = []
            for row in ws.iter_rows(min_row=linha_header + 1, values_only=True):
                mat = _txt(row[mapa["matricula"]] if mapa["matricula"] < len(row) else None)
                mat = mat.replace(".0", "")
                if not mat.isdigit():
                    continue

                def col(campo, padrao=0.0):
                    i = mapa.get(campo)
                    return _num(row[i], padrao) if i is not None and i < len(row) else padrao

                linhas.append({"mat": int(mat), "salario": col("salario"), "vt": col("vt"),
                               "va": col("va"), "faltas_dias": col("faltas_dias"),
                               "faltas_horas": col("faltas_horas"),
                               "saldo_livre": col("saldo_livre"), "acerto": col("acerto"),
                               "premiacoes": col("premiacoes"),
                               "total_planilha": col("total_planilha"),
                               "provisoes_planilha": col("provisoes_planilha")})

            if not linhas:
                self.stdout.write(self.style.WARNING(f"  {aba}: sem linhas — pulando"))
                continue

            if dry:
                self.stdout.write(
                    f"  [dry-run] {mes:02d}/{ano} ({aba}): {len(linhas)} pessoas, "
                    f"{dias_mes} dias / {dias_uteis} úteis")
                continue

            with transaction.atomic():
                comp, _ = DpCompetencia.objects.get_or_create(
                    ano=ano, mes=mes, defaults={"dias_mes": dias_mes, "dias_uteis": dias_uteis})
                if comp.status == "fechada":
                    self.stdout.write(self.style.WARNING(
                        f"  {mes:02d}/{ano}: já FECHADA — não mexo"))
                    continue
                comp.dias_mes, comp.dias_uteis = dias_mes, dias_uteis
                comp.save()

                fiscal = tabela_fiscal_para(ano, mes)
                DpFolhaItem.objects.filter(competencia=comp).delete()
                itens, ausentes, divergentes = [], [], 0
                soma_total = soma_custo = soma_prov = 0.0
                soma_total_pl = soma_prov_pl = 0.0

                for l in linhas:
                    colab = por_matricula.get(l["mat"])
                    if not colab:
                        ausentes.append(l["mat"])
                        continue
                    # override só quando o mês DIFERE do cadastro de hoje — é
                    # exatamente o que o campo significa (ajuste pontual do mês)
                    def ov(valor, atual):
                        return valor if abs(valor - _num(atual)) > 0.01 else None

                    lanc, _ = DpLancamento.objects.update_or_create(
                        competencia=comp, colaborador=colab,
                        defaults={
                            "faltas_dias": l["faltas_dias"], "faltas_horas": l["faltas_horas"],
                            "premiacoes": l["premiacoes"], "acerto_contabil": l["acerto"],
                            "salario_override": ov(l["salario"], colab.salario_bruto),
                            "vt_override": ov(l["vt"], colab.vt),
                            "va_override": ov(l["va"], colab.va),
                            "saldo_livre_override": ov(l["saldo_livre"], colab.saldo_livre),
                        })
                    d = calcular_item(colab, lanc, comp, fiscal)
                    itens.append(DpFolhaItem(competencia=comp, colaborador=colab, **d))
                    soma_total += d["total_pagar"]
                    soma_custo += d["custo_total"]
                    soma_prov += d["custo_provisoes"]
                    soma_total_pl += l["total_planilha"]
                    soma_prov_pl += l["provisoes_planilha"]
                    if l["total_planilha"] and abs(d["total_pagar"] - l["total_planilha"]) > 0.05:
                        divergentes += 1

                DpFolhaItem.objects.bulk_create(itens, batch_size=500)

            dif_total = soma_total - soma_total_pl
            dif_prov = soma_prov - soma_prov_pl
            self.stdout.write(
                f"  {mes:02d}/{ano} ({aba}): {len(itens)} itens · "
                f"a pagar R$ {soma_total:,.2f} (planilha R$ {soma_total_pl:,.2f}, "
                f"dif {dif_total:+,.2f}) · provisões R$ {soma_prov:,.2f} "
                f"(planilha R$ {soma_prov_pl:,.2f}, dif {dif_prov:+,.2f}) · "
                f"custo total R$ {soma_custo:,.2f}")
            if ausentes:
                self.stdout.write(self.style.WARNING(
                    f"     {len(ausentes)} matrícula(s) da aba sem cadastro: {ausentes[:8]}"))
            if divergentes:
                self.stdout.write(self.style.WARNING(
                    f"     {divergentes} linha(s) divergem do TOTAL da planilha "
                    f"(acima de 5 centavos)"))
            resumo.append((mes, len(itens), soma_custo, dif_total, dif_prov))

        wb.close()

        # ── fechamento ──
        fechar_ate = o["fechar_ate"]
        if fechar_ate and not dry:
            from financeiro.models_estrutura import congelar_competencia
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(
                f"Fechando competências até {fechar_ate:02d}/{ano}"))
            self.stdout.write(self.style.WARNING(
                "ATENÇÃO: isto é CARGA HISTÓRICA, não aprovação. O fechamento normal\n"
                "exige 4 olhos (quem envia à revisão não pode aprovar) e essas\n"
                "competências não passaram por ele — vieram prontas da planilha.\n"
                "A trilha de auditoria registra isso explicitamente."))
            for mes in range(o["de"], min(fechar_ate, o["ate"]) + 1):
                comp = DpCompetencia.objects.filter(ano=ano, mes=mes).first()
                if not comp or comp.status == "fechada":
                    continue
                comp.status = "fechada"
                comp.fechada_por = USUARIO
                comp.fechada_em = timezone.now()
                comp.save()
                n_p, n_a = congelar_competencia(comp)
                DpAuditLog.objects.create(
                    usuario=USUARIO, acao="fechar_competencia",
                    entidade="dp_competencia", entidade_id=str(comp.id),
                    depois={"fechada_por": USUARIO,
                            "foto_enquadramentos": n_p, "foto_alocacoes": n_a,
                            "observacao": "Carga histórica da planilha do DP — NÃO passou "
                                          "pelo fluxo de 4 olhos."})
                self.stdout.write(self.style.SUCCESS(
                    f"  {mes:02d}/{ano} fechada e congelada ({n_p} pessoas, {n_a} alocações)"))

        if resumo and not dry:
            self.stdout.write("")
            total_custo = sum(r[2] for r in resumo)
            pior_total = max((abs(r[3]) for r in resumo), default=0)
            pior_custo = max((abs(r[4]) for r in resumo), default=0)
            self.stdout.write(self.style.SUCCESS(
                f"{len(resumo)} competência(s) carregadas · custo somado "
                f"R$ {total_custo:,.2f} · maior divergência vs planilha: "
                f"a pagar R$ {pior_total:,.2f}, provisões R$ {pior_custo:,.2f}"))
