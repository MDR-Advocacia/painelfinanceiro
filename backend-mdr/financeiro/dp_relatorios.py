# DP — F4: dashboard consolidado + RELATÓRIOS exportáveis TIMBRADOS (logo MDR).
#   • GET /dp/dashboard/                              KPIs + série mensal + regime
#   • GET /dp/competencias/<id>/relatorio/?tipo=folha|rateio&formato=excel|pdf
#   • GET /dp/relatorio-quadro/                       Excel do quadro de pessoal
# Excel: openpyxl com cabeçalho timbrado. PDF: reportlab (rateio — o relatório
# que vai pro fechamento/diretoria).
import os
from datetime import date, datetime
from io import BytesIO

from django.db.models import Count, Q, Sum

from .dp_folha import MES_NOMES
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from .dp_audit import humanizar
from .dp_escopo import filtrar_colaboradores, filtrar_folha
from .dp_views import filtrar_quadro
from .models import DpCentroCusto, DpColaborador, DpCompetencia, DpEvento, DpFolhaItem
from .views import modulo_permission

_PERM = [modulo_permission(read_any=["pessoal"], write="pessoal")]
_LOGO = os.path.join(os.path.dirname(__file__), "assets", "logo-mdr.png")

AZUL = "1E7BFF"
NAVY = "0A1940"


# ─────────────────────────────── DASHBOARD ───────────────────────────────

@api_view(["GET"])
@permission_classes(_PERM)
def dp_dashboard(request):
    hoje = date.today()
    # ── ESCOPO: quais competências alimentam os KPIs ────────────────────────
    # Default é FECHADAS. A folha aberta é um retrato pela metade — em agosto,
    # antes do fechamento, faltam faltas, prêmios e acertos do mês, e o painel
    # mostrava esse número incompleto como se fosse o custo do escritório.
    # Quem quiser ver o mês em curso pede escopo=todas explicitamente.
    escopo = (request.query_params.get("escopo") or "fechadas").lower()
    if escopo not in ("fechadas", "todas"):
        escopo = "fechadas"
    comps_qs = DpCompetencia.objects.order_by("ano", "mes")
    if escopo == "fechadas":
        comps_qs = comps_qs.filter(status="fechada")
        # se nenhuma fechou ainda, cair pra todas em vez de devolver painel vazio
        if not comps_qs.exists():
            comps_qs = DpCompetencia.objects.order_by("ano", "mes")
            escopo = "todas"
    ativos = filtrar_colaboradores(DpColaborador.objects.filter(status="ativo"), request.user)
    headcount = ativos.count()
    por_regime = dict(ativos.values_list("regime").annotate(n=Count("id")))

    # admissões/desligamentos por mês (últimos 12 meses, via eventos).
    # TRANSFERÊNCIA DE CONTRATO fica de fora: quando alguém é efetivado, a casa
    # encerra a matrícula antiga e abre uma nova (a numeração é por regime), o
    # que gera um desligamento e uma admissão da MESMA pessoa. Contar isso
    # inflaria o turnover com quem nunca saiu do escritório.
    from .models import ids_em_transferencia
    saidas_transf, entradas_transf = ids_em_transferencia()
    serie_mov = []
    for i in range(11, -1, -1):
        a, m = hoje.year, hoje.month - i
        while m <= 0:
            m += 12
            a -= 1
        ini, fim = date(a, m, 1), date(a + (m == 12), (m % 12) + 1, 1)
        adm = (DpEvento.objects.filter(tipo__in=["admissao", "importacao"],
                                       data_efeito__gte=ini, data_efeito__lt=fim)
               .exclude(colaborador_id__in=entradas_transf).count())
        des = (DpEvento.objects.filter(tipo="desligamento",
                                       data_efeito__gte=ini, data_efeito__lt=fim)
               .exclude(colaborador_id__in=saidas_transf).count())
        # HEADCOUNT MEDIO DO MES, nao o de hoje. Dividir o desligamento de
        # janeiro pelo quadro de agosto compara epocas diferentes e o indice
        # anda sozinho toda vez que o escritorio cresce.
        hc_fim = (DpColaborador.objects
                  .filter(data_admissao__lt=fim)
                  .exclude(data_demissao__lt=fim)
                  .exclude(id__in=entradas_transf).count())
        hc_ini = hc_fim - adm + des
        hc_medio = (hc_ini + hc_fim) / 2 or 0
        # DOIS indices, porque medem coisas diferentes:
        #  · desligamento  = perda de gente (o que interessa pra retencao)
        #  · classico      = rotatividade (media de entradas e saidas, Chiavenato)
        t_deslig = round(des / hc_medio * 100, 2) if hc_medio else 0.0
        t_classico = round(((adm + des) / 2) / hc_medio * 100, 2) if hc_medio else 0.0
        serie_mov.append({"mes": f"{a}-{m:02d}", "admissoes": adm, "desligamentos": des,
                          "headcount_medio": round(hc_medio, 1),
                          "turnover": t_deslig, "turnover_classico": t_classico})

    # série mensal COMPLETA por competência (espelha a aba de evolução da planilha)
    serie_custo = []
    fgts_acum = multa_acum = 0.0
    mov_por_mes = {m["mes"]: m for m in serie_mov}
    for comp in comps_qs[:24]:
        itens = filtrar_folha(comp.itens.all(), request.user)
        ag = itens.aggregate(pagar=Sum("total_pagar"), prov=Sum("custo_provisoes"),
                             pat=Sum("inss_patronal"), custo=Sum("custo_total"),
                             fgts=Sum("fgts_mensal"), multa=Sum("multa_fgts_mensal"),
                             sf=Sum("salario_familia"))
        ag_clt = itens.filter(regime="clt").aggregate(custo=Sum("custo_total"))
        hc = itens.count()
        chave = f"{comp.ano}-{comp.mes:02d}"
        mov = mov_por_mes.get(chave, {"admissoes": 0, "desligamentos": 0})
        fgts_acum += round(ag["fgts"] or 0, 2)
        multa_acum += round(ag["multa"] or 0, 2)
        serie_custo.append({
            "mes": chave, "status": comp.status, "headcount": hc,
            "admissoes": mov["admissoes"], "desligamentos": mov["desligamentos"],
            "turnover": round(mov["desligamentos"] / hc * 100, 2) if hc else 0.0,
            # folha = parcela de CUSTO (pago menos o salário-família, que o INSS
            # reembolsa) para que folha + provisoes + patronal feche com custo_total
            "folha": round((ag["pagar"] or 0) - (ag["sf"] or 0), 2),
            "a_pagar": round(ag["pagar"] or 0, 2),
            "salario_familia": round(ag["sf"] or 0, 2),
            "provisoes": round(ag["prov"] or 0, 2),
            "patronal": round(ag["pat"] or 0, 2), "custo_total": round(ag["custo"] or 0, 2),
            "custo_clt": round(ag_clt["custo"] or 0, 2),
            "fgts": round(ag["fgts"] or 0, 2), "multa_fgts": round(ag["multa"] or 0, 2),
            "fgts_acumulado": round(fgts_acum, 2), "multa_fgts_acumulada": round(multa_acum, 2),
        })

    ult = serie_custo[-1] if serie_custo else None
    mov_mes = serie_mov[-1] if serie_mov else {"admissoes": 0, "desligamentos": 0}
    turnover = mov_mes.get("turnover", 0.0)

    # ── TURNOVER ACUMULADO POR JANELA ──────────────────────────────────────
    # Nao e' a media das taxas mensais: soma-se gente e divide-se por gente.
    # A media aritmetica das taxas daria peso igual a um mes de 200 pessoas e a
    # um de 30, o que distorce quando o quadro cresce no meio do periodo.
    def _janela(n):
        janela = serie_mov[-n:] if n else serie_mov
        if not janela:
            return None
        adm = sum(x["admissoes"] for x in janela)
        des = sum(x["desligamentos"] for x in janela)
        hc = sum(x["headcount_medio"] for x in janela) / len(janela)
        return {
            "meses": len(janela), "admissoes": adm, "desligamentos": des,
            "headcount_medio": round(hc, 1),
            "turnover": round(des / hc * 100, 2) if hc else 0.0,
            "turnover_classico": round(((adm + des) / 2) / hc * 100, 2) if hc else 0.0,
            # media MENSAL da janela — comparavel com a taxa de um mes isolado
            "turnover_mensal_medio": round(des / hc * 100 / len(janela), 2) if hc else 0.0,
        }

    turnover_janelas = {r: _janela(n) for r, n in
                        (("3m", 3), ("6m", 6), ("12m", 12))}

    # ── análises extras (mais profundidade no painel) ──
    from django.db.models import Avg

    por_unidade = [{"nome": r["unidade"] or "(sem unidade)", "quantidade": r["n"]}
                   for r in ativos.values("unidade").annotate(n=Count("id")).order_by("-n")]
    por_area = [{"nome": r["area"] or "(sem área)", "quantidade": r["n"]}
                for r in ativos.values("area").annotate(n=Count("id")).order_by("-n")]
    por_cc_qtd = [{"nome": r["centro_custo__nome"] or "(sem centro de custo)",
                   "quantidade": r["n"],
                   "salario_medio": round(r["media"] or 0, 2)}
                  for r in ativos.values("centro_custo__nome")
                  .annotate(n=Count("id"), media=Avg("salario_bruto")).order_by("-n")]

    # custo por centro de custo (da última competência calculada)
    custo_por_cc, custo_medio_pessoa, participacao = [], 0.0, {}
    comp_ult = comps_qs.order_by("-ano", "-mes").first()
    if comp_ult:
        itens_ult = filtrar_folha(comp_ult.itens.all(), request.user)
        custo_por_cc = [{"nome": r["centro_custo_nome"], "quantidade": r["n"],
                         "custo": round(r["custo"] or 0, 2),
                         "custo_medio": round((r["custo"] or 0) / r["n"], 2) if r["n"] else 0}
                        for r in itens_ult.values("centro_custo_nome")
                        .annotate(n=Count("id"), custo=Sum("custo_total")).order_by("-custo")]
        ag_ult = itens_ult.aggregate(pagar=Sum("total_pagar"), prov=Sum("custo_provisoes"),
                                     pat=Sum("inss_patronal"), custo=Sum("custo_total"),
                                     sf=Sum("salario_familia"), n=Count("id"))
        total_ult = ag_ult["custo"] or 0
        custo_medio_pessoa = round(total_ult / ag_ult["n"], 2) if ag_ult["n"] else 0
        if total_ult:
            participacao = {
                # numerador SEM o salário-família: senão os três pedaços somariam
                # mais de 100% do custo, porque o denominador não o inclui
                "folha": round(((ag_ult["pagar"] or 0) - (ag_ult["sf"] or 0)) / total_ult * 100, 1),
                "provisoes": round((ag_ult["prov"] or 0) / total_ult * 100, 1),
                "patronal": round((ag_ult["pat"] or 0) / total_ult * 100, 1),
            }

    # custo médio POR CARGO (da última competência)
    custo_por_cargo = []
    if comp_ult:
        for r in (filtrar_folha(comp_ult.itens.all(), request.user)
                  .exclude(cargo_nome="").values("cargo_nome", "regime")
                  .annotate(n=Count("id"), custo=Sum("custo_total"),
                            salario=Sum("salario_bruto")).order_by("-custo")):
            custo_por_cargo.append({
                "cargo": r["cargo_nome"], "regime": r["regime"], "quantidade": r["n"],
                "custo": round(r["custo"] or 0, 2),
                "custo_medio": round((r["custo"] or 0) / r["n"], 2) if r["n"] else 0,
                "salario_medio": round((r["salario"] or 0) / r["n"], 2) if r["n"] else 0,
            })

    # custo médio por tipo de contrato (da última competência)
    custo_por_regime = []
    if comp_ult:
        for r in (filtrar_folha(comp_ult.itens.all(), request.user).values("regime")
                  .annotate(n=Count("id"), custo=Sum("custo_total")).order_by("-custo")):
            custo_por_regime.append({
                "regime": r["regime"], "quantidade": r["n"],
                "custo": round(r["custo"] or 0, 2),
                "custo_medio": round((r["custo"] or 0) / r["n"], 2) if r["n"] else 0})

    # variação do custo vs mês anterior
    variacao = None
    if len(serie_custo) >= 2:
        atual_v, ant_v = serie_custo[-1]["custo_total"], serie_custo[-2]["custo_total"]
        if ant_v:
            variacao = {"percent": round((atual_v - ant_v) / ant_v * 100, 1),
                        "valor": round(atual_v - ant_v, 2)}

    # tempo de casa (faixas) — ativos
    from datetime import timedelta
    faixas_casa = {"Até 6 meses": 0, "6 a 12 meses": 0, "1 a 2 anos": 0, "Mais de 2 anos": 0}
    for c in ativos.only("data_admissao", "data_entrada"):
        ref = c.data_admissao or c.data_entrada
        if not ref:
            continue
        dias = (hoje - ref).days
        if dias <= 182:
            faixas_casa["Até 6 meses"] += 1
        elif dias <= 365:
            faixas_casa["6 a 12 meses"] += 1
        elif dias <= 730:
            faixas_casa["1 a 2 anos"] += 1
        else:
            faixas_casa["Mais de 2 anos"] += 1

    # alertas operacionais
    alertas = []
    sem_cc = ativos.filter(centro_custo__isnull=True).count()
    sem_cargo = ativos.filter(cargo__isnull=True).count()
    sem_salario = ativos.filter(salario_bruto__lte=0).count()
    if sem_cargo:
        alertas.append({"tipo": "atencao", "texto": f"{sem_cargo} colaborador(es) ativo(s) sem cargo definido"})
    if sem_salario:
        alertas.append({"tipo": "critico", "texto": f"{sem_salario} colaborador(es) ativo(s) sem salário cadastrado"})
    if sem_cc:
        alertas.append({"tipo": "critico", "texto": f"{sem_cc} colaborador(es) sem centro de custo"})
    comp_aberta = DpCompetencia.objects.filter(status="aberta").order_by("-ano", "-mes").first()
    if comp_aberta:
        alertas.append({"tipo": "info",
                        "texto": f"Competência {comp_aberta.mes:02d}/{comp_aberta.ano} está aberta (ainda não fechada)"})

    # ── PREMIAÇÕES: quanto cada centro de custo pagou de prêmio ──
    premiacoes = _premiacoes_por_cc(request)

    return Response({
        "premiacoes": premiacoes,
        "headcount": headcount, "por_regime": por_regime,
        "admissoes_mes": mov_mes["admissoes"], "desligamentos_mes": mov_mes["desligamentos"],
        "turnover_mes": turnover,
        "turnover_classico_mes": mov_mes.get("turnover_classico", 0.0),
        "turnover_janelas": turnover_janelas,
        "custo_competencia": ult, "serie_mov": serie_mov, "serie_custo": serie_custo,
        # de onde vieram os numeros: o painel precisa DIZER isso, senao o
        # operador compara com a folha da tela ao lado e nao entende a diferenca
        "escopo": escopo,
        "competencia_base": ({"ano": comp_ult.ano, "mes": comp_ult.mes,
                              # mes_nome e' campo do SERIALIZER, nao do model —
                              # ler do objeto direto dava 500 no dashboard
                              "mes_nome": MES_NOMES[comp_ult.mes],
                              "status": comp_ult.status}
                             if comp_ult else None),
        "competencias_fechadas": DpCompetencia.objects.filter(status="fechada").count(),
        # extras
        "por_unidade": por_unidade, "por_area": por_area, "por_cc_qtd": por_cc_qtd,
        "custo_por_cc": custo_por_cc, "custo_por_regime": custo_por_regime,
        "custo_por_cargo": custo_por_cargo,
        "custo_medio_pessoa": custo_medio_pessoa, "participacao": participacao,
        "variacao_custo": variacao,
        "tempo_casa": [{"faixa": k, "quantidade": v} for k, v in faixas_casa.items()],
        "alertas": alertas,
        "folha_media_salario": round(ativos.aggregate(m=Avg("salario_bruto"))["m"] or 0, 2),
    })


def _premiacoes_por_cc(request) -> dict:
    """Premiações pagas: no mês da última competência, no ano e mês a mês.

    Premiação é o único valor 100% discricionário da folha — por isso ganha
    recorte próprio: quem está distribuindo, quanto e para quantas pessoas.
    """
    comps = list(DpCompetencia.objects.order_by("ano", "mes"))
    if not comps:
        return {"mes": [], "ano": [], "serie": [], "top": [],
                "total_mes": 0.0, "total_ano": 0.0, "competencia": None, "ano_ref": None}

    ult = comps[-1]
    ano_ref = ult.ano

    # nome do CC → id, pro clique na tabela abrir o quadro já filtrado
    ids_cc = {c.nome: str(c.id) for c in DpCentroCusto.objects.all()}

    def por_cc(qs):
        linhas = [{"nome": r["centro_custo_nome"] or "(sem centro de custo)",
                   "cc_id": ids_cc.get(r["centro_custo_nome"] or ""),
                   "valor": round(r["v"] or 0, 2),
                   "pessoas": r["n"],
                   "media": round((r["v"] or 0) / r["n"], 2) if r["n"] else 0.0}
                  for r in (qs.filter(premiacoes__gt=0).values("centro_custo_nome")
                            .annotate(v=Sum("premiacoes"), n=Count("colaborador_id", distinct=True))
                            .order_by("-v"))]
        total = sum(x["valor"] for x in linhas)
        for x in linhas:
            x["percentual"] = round(x["valor"] / total * 100, 1) if total else 0.0
        return linhas, round(total, 2)

    itens_mes = filtrar_folha(DpFolhaItem.objects.filter(competencia=ult), request.user)
    itens_ano = filtrar_folha(
        DpFolhaItem.objects.filter(competencia__ano=ano_ref), request.user)
    linhas_mes, total_mes = por_cc(itens_mes)
    linhas_ano, total_ano = por_cc(itens_ano)

    serie = []
    for c in comps[-12:]:
        ag = filtrar_folha(DpFolhaItem.objects.filter(competencia=c), request.user).aggregate(
            v=Sum("premiacoes"), n=Count("id", filter=Q(premiacoes__gt=0)))
        serie.append({"mes": f"{c.ano}-{c.mes:02d}", "valor": round(ag["v"] or 0, 2),
                      "pessoas": ag["n"] or 0})

    top = [{"nome": r["nome"], "centro_custo": r["centro_custo_nome"],
            "valor": round(r["v"] or 0, 2), "meses": r["n"]}
           for r in (itens_ano.filter(premiacoes__gt=0)
                     .values("nome", "centro_custo_nome")
                     .annotate(v=Sum("premiacoes"), n=Count("id")).order_by("-v")[:10])]

    return {
        "competencia": f"{ult.mes:02d}/{ult.ano}", "ano_ref": ano_ref,
        "mes": linhas_mes, "ano": linhas_ano, "serie": serie, "top": top,
        "total_mes": total_mes, "total_ano": total_ano,
        "pessoas_mes": itens_mes.filter(premiacoes__gt=0).count(),
        "pessoas_ano": itens_ano.filter(premiacoes__gt=0).values("colaborador_id").distinct().count(),
    }


# ─────────────────────────────── HELPERS EXCEL ───────────────────────────────

def _wb_timbrado(titulo: str, subtitulo: str, usuario: str):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    if os.path.exists(_LOGO):
        img = XLImage(_LOGO)
        # proporção da logo (~450×160) reduzida pro cabeçalho
        img.width, img.height = 132, 47
        ws.add_image(img, "A1")
    ws.merge_cells("C1:H1")
    ws["C1"] = titulo
    ws["C1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws.merge_cells("C2:H2")
    ws["C2"] = subtitulo
    ws["C2"].font = Font(name="Calibri", size=10, color="666666")
    ws.merge_cells("C3:H3")
    ws["C3"] = f"Gerado por {usuario} em {datetime.now().strftime('%d/%m/%Y %H:%M')} — MDR Advocacia · Painel Financeiro"
    ws["C3"].font = Font(name="Calibri", size=8, italic=True, color="999999")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 14
    return wb, ws


def _tabela(ws, linha0: int, headers: list, rows: list, larguras: list = None, money_cols: set = None):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor=NAVY)
    fina = Side(style="thin", color="DDDDDD")
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)
    money = money_cols or set()

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=linha0, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda
    for i, row in enumerate(rows, start=linha0 + 1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(size=9)
            c.border = borda
            if j in money:
                c.number_format = 'R$ #,##0.00'
    if larguras:
        for j, w in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.auto_filter.ref = f"A{linha0}:{get_column_letter(len(headers))}{linha0 + len(rows)}"
    ws.freeze_panes = ws.cell(row=linha0 + 1, column=1)


def _resposta_excel(wb, nome: str) -> HttpResponse:
    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{nome}"'
    resp["Access-Control-Expose-Headers"] = "Content-Disposition"
    return resp


def _quem(request):
    u = request.user
    return (u.email or u.username) if u and u.is_authenticated else "?"


# ─────────────────────────────── RELATÓRIOS ───────────────────────────────

def _linhas_extrato(it):
    """Verbas de UM colaborador no formato do EXTRATO MENSAL da contabilidade.

    Os códigos são os MESMOS do sistema da Montenegro (8781 dias normais, 812
    INSS férias…) de propósito: o DP confere lado a lado, e código igual elimina
    o "qual linha daqui é qual linha de lá". Devolve (proventos, descontos,
    informativas) — informativa é o FGTS, que não entra em nenhum dos totais.
    """
    P, D = [], []

    def p(cod, desc, ref, valor):
        if valor:
            P.append((cod, desc, ref, round(valor, 2)))

    def d(cod, desc, ref, valor):
        if valor:
            D.append((cod, desc, ref, round(valor, 2)))

    estag = it.regime == "estagiario"
    dias_norm = 30 - (it.ferias_dias or 0)
    if estag:
        p("8797", "DIAS BOLSA AUXILIO", f"{dias_norm:g}", it.salario_com_faltas)
    else:
        p("8781", "DIAS NORMAIS", f"{dias_norm:g}", it.salario_com_faltas)
    p("8783", "DIAS FERIAS", f"{it.ferias_dias:g}" if it.ferias_dias else "",
      it.ferias_valor)
    p("931", "1/3 DAS FERIAS", "33,33", it.ferias_terco)
    p("8800", "ABONO PECUNIARIO + 1/3", "", it.ferias_abono)
    p("995", "SALARIO FAMILIA", f"{it.salario_familia_cotas or ''}", it.salario_familia)
    p("201", "AUXILIO TRANSPORTE (BENEFICIO)", "", it.vt_com_faltas)
    p("202", "AUXILIO ALIMENTACAO (BENEFICIO)", "", it.va_com_faltas)
    p("300", "SALDO LIVRE", "", it.saldo_livre)
    p("310", "PREMIACOES", "", it.premiacoes)
    p("320", "ACERTO CONTABIL", "", max(it.acerto_contabil or 0, 0))
    p("330", "13o SALARIO (PARCELA PAGA)", "", it.decimo_terceiro_pago)

    d("998", "I.N.S.S.", "", it.inss_salario if it.inss_ferias else it.desc_inss)
    d("812", "INSS FERIAS", "", it.inss_ferias)
    d("821", "INSS DIFERENCA FERIAS", "", it.inss_dif_ferias)
    d("48", "VALE TRANSPORTE", "6,00", it.desc_vt)
    d("11", "I.R.R.F.", "", it.desc_irrf)
    d("937", "ADIANTAMENTO DE FERIAS", "", it.adiantamento_ferias)
    d("9750", "EMPRESTIMO CONSIGNADO", "", it.desconto_consignado)
    d("999", "OUTROS DESCONTOS", "", it.outros_descontos)
    d("321", "ACERTO CONTABIL (DESCONTO)", "", -min(it.acerto_contabil or 0, 0))
    return P, D


def _pdf_extrato_mensal(comp, itens, usuario):
    """EXTRATO MENSAL em PDF — um bloco por colaborador, fechado com
    Proventos/Descontos/Líquido e as bases, igual ao relatório da contabilidade.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm,
                            bottomMargin=12 * mm, leftMargin=12 * mm,
                            rightMargin=12 * mm)
    st_t = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=13)
    st_s = ParagraphStyle("s", fontName="Helvetica", fontSize=8,
                          textColor=colors.HexColor("#555555"))
    st_n = ParagraphStyle("n", fontName="Helvetica-Bold", fontSize=9.5)
    el = [Paragraph(f"EXTRATO MENSAL — {comp.mes:02d}/{comp.ano}", st_t),
          Paragraph(f"{len(itens)} colaborador(es) · gerado por {usuario} · "
                    f"competência {comp.get_status_display().lower()}", st_s),
          Spacer(1, 4 * mm)]

    cinza = colors.HexColor("#F2F4F7")
    for it in itens:
        P, D = _linhas_extrato(it)
        el.append(Paragraph(
            f"{it.matricula} · {it.nome} — {REG_LABEL.get(it.regime, it.regime)}"
            + (f" · {it.centro_custo_nome}" if it.centro_custo_nome else ""), st_n))
        linhas = [["Cód", "Provento", "Ref", "Valor", "Cód", "Desconto", "Ref", "Valor"]]
        for i in range(max(len(P), len(D))):
            pl = P[i] if i < len(P) else ("", "", "", "")
            dl = D[i] if i < len(D) else ("", "", "", "")
            linhas.append([pl[0], pl[1], pl[2], _brl(pl[3]) if pl[3] else "",
                           dl[0], dl[1], dl[2], _brl(dl[3]) if dl[3] else ""])
        linhas.append(["", "TOTAL PROVENTOS", "", _brl(it.total_proventos),
                       "", "TOTAL DESCONTOS", "", _brl(it.total_descontos)])
        linhas.append(["", f"Base INSS: {_brl(it.base_inss)}", "",
                       f"FGTS: {_brl(it.fgts)}",
                       "", "LÍQUIDO EM CONTA", "", _brl(it.liquido_em_conta)])
        t = Table(linhas, colWidths=[11 * mm, 52 * mm, 10 * mm, 20 * mm,
                                     11 * mm, 52 * mm, 10 * mm, 20 * mm])
        t.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), "Helvetica", 7.2),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 7.2),
            ("FONT", (0, -2), (-1, -1), "Helvetica-Bold", 7.2),
            ("BACKGROUND", (0, 0), (-1, 0), cinza),
            ("BACKGROUND", (0, -2), (-1, -1), cinza),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("ALIGN", (7, 0), (7, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 1.4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.4),
        ]))
        el.append(t)
        el.append(Spacer(1, 4 * mm))

    doc.build(el)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = (
        f'attachment; filename="extrato_mensal_{comp.ano}_{comp.mes:02d}.pdf"')
    return resp


@api_view(["GET"])
@permission_classes(_PERM)
def dp_relatorio_competencia(request, pk):
    comp = DpCompetencia.objects.filter(pk=pk).first()
    if not comp:
        return Response({"detail": "Competência não encontrada."}, status=404)
    tipo = request.query_params.get("tipo", "folha")
    formato = request.query_params.get("formato", "excel")
    rotulo = f"{comp.mes:02d}/{comp.ano}"

    if tipo == "extrato":
        itens = comp.itens.select_related("colaborador").order_by("nome")
        colab_id = request.query_params.get("colaborador")
        if colab_id:
            itens = itens.filter(colaborador_id=colab_id)
        itens = list(itens)
        if not itens:
            return Response({"detail": "Nenhum item na competência para esse filtro."},
                            status=404)
        return _pdf_extrato_mensal(comp, itens, _quem(request))

    if tipo == "rateio":
        # usa o mesmo resumo detalhado da tela (espelho da aba CC da planilha)
        from .dp_folha import DpCompetenciaViewSet as _VS
        vs = _VS()
        vs.request = request
        dados = vs.rateio(request, pk=str(comp.pk)).data
        det, tot_d = dados["linhas"], dados["totais"]

        rows = [[l["centro_custo_nome"], l["headcount"], round(l["folha"] or 0, 2),
                 round(l["provisoes"] or 0, 2), round(l["patronal"] or 0, 2),
                 round(l["custo"] or 0, 2)] for l in det]
        tot = ["TOTAL", tot_d["headcount"], tot_d["folha"], tot_d["provisoes"],
               tot_d["patronal"], tot_d["custo"]]

        if formato == "pdf":
            return _pdf_rateio(comp, rows, tot, _quem(request), rotulo)

        wb, ws = _wb_timbrado("Resumo do Centro de Custo por Setor",
                              f"Competência {rotulo} · situação: {comp.status}", _quem(request))
        _tabela(ws, 5, ["Centro de Custo", "Núcleo", "Colaboradores", "Salários (R$)",
                        "VT (R$)", "VA (R$)", "Saldo livre (R$)", "Premiações (R$)",
                        "Acertos (R$)", "INSS patronal (R$)", "Custo mensal (R$)", "% do custo",
                        "13º (R$)", "Férias (R$)", "1/3 férias (R$)", "FGTS (R$)",
                        "Multa FGTS (R$)", "Recesso (R$)", "Total provisões (R$)"],
                [[l["centro_custo_nome"], l["nucleo"], l["headcount"], l["salarios"], l["vt"],
                  l["va"], l["saldo_livre"], l["premios"], l["acertos"], l["patronal"],
                  l["custo"], l["percentual"] / 100, l["decimo"], l["ferias"], l["terco"],
                  l["fgts"], l["multa_fgts"], l["recesso"], l["provisoes"]] for l in det]
                + [["TOTAL", "", tot_d["headcount"], tot_d["salarios"], tot_d["vt"], tot_d["va"],
                    tot_d["saldo_livre"], tot_d["premios"], tot_d["acertos"], tot_d["patronal"],
                    tot_d["custo"], 1, tot_d["decimo"], tot_d["ferias"], tot_d["terco"],
                    tot_d["fgts"], tot_d["multa_fgts"], tot_d["recesso"], tot_d["provisoes"]]],
                larguras=[30, 14, 13, 15, 12, 12, 15, 14, 13, 17, 16, 11, 12, 13, 14, 12, 15, 13, 17],
                money_cols={4, 5, 6, 7, 8, 9, 10, 11, 13, 14, 15, 16, 17, 18, 19})
        ws.cell(row=1, column=12).number_format = "0.0%"
        wsn = wb.create_sheet("Por núcleo")
        _tabela(wsn, 1, ["Núcleo", "Centros de custo", "Colaboradores", "Folha (R$)",
                         "Provisões (R$)", "INSS patronal (R$)", "Custo total (R$)", "% do custo",
                         "Sal. família (R$)"],
                [[n["nucleo"], n["centros"], n["headcount"], n["folha"], n["provisoes"],
                  n["patronal"], n["custo"], n["percentual"] / 100,
                  n.get("salario_familia", 0)] for n in dados["nucleos"]],
                larguras=[26, 17, 14, 16, 16, 18, 17, 11, 18], money_cols={4, 5, 6, 7, 9})
        return _resposta_excel(wb, f"resumo_cc_{comp.ano}_{comp.mes:02d}.xlsx")

    # ── RESCISOES DA COMPETENCIA ───────────────────────────────────────────
    # O relatorio mostrava a folha do mes e calava sobre quem foi desligado —
    # justamente o evento de maior impacto financeiro do mes. Traz a marca na
    # linha da pessoa e as verbas rescisorias numa aba/bloco proprio.
    from .models import DpRescisao
    _ini = date(comp.ano, comp.mes, 1)
    _fim = date(comp.ano + (comp.mes == 12), (comp.mes % 12) + 1, 1)
    rescisoes = list(DpRescisao.objects.filter(
        data_desligamento__gte=_ini, data_desligamento__lt=_fim)
        .select_related("colaborador").order_by("data_desligamento"))
    resc_por_colab = {r.colaborador_id: r for r in rescisoes}
    resc_total = round(sum(r.liquido for r in rescisoes), 2)

    def _marca(it):
        """Sinaliza a rescisao na linha da pessoa."""
        r = resc_por_colab.get(it.colaborador_id)
        if not r:
            return ""
        return f"RESCISÃO {r.data_desligamento.strftime('%d/%m')} · {r.get_tipo_display()}"

    # folha analítica
    if formato == "pdf":
        rows = [[it.matricula, it.nome[:30], REG_LABEL.get(it.regime, it.regime),
                 it.centro_custo_nome[:20], _brl(it.salario_bruto), _brl(it.desc_inss),
                 _brl(it.total_pagar), _brl(it.custo_total), _marca(it)]
                for it in comp.itens.all()]
        tot = ["TOTAL", "", "", "", "", "",
               _brl(sum(i.total_pagar for i in comp.itens.all())),
               _brl(sum(i.custo_total for i in comp.itens.all())), ""]
        # as verbas rescisorias fecham o relatorio, com o liquido de cada saida
        for r in rescisoes:
            rows.append(["", f"↳ rescisão: {r.colaborador.nome[:26]}",
                         r.get_tipo_display()[:18],
                         f"aviso {r.aviso_dias}d", _brl(r.proventos),
                         _brl(r.total_descontos), _brl(r.liquido), "",
                         r.data_desligamento.strftime("%d/%m/%Y")])
        if rescisoes:
            rows.append(["", f"TOTAL DAS RESCISÕES ({len(rescisoes)})", "", "", "", "",
                         _brl(resc_total), "", ""])
        return _pdf_generico(f"Folha Analítica — {rotulo}",
                             f"{comp.itens.count()} colaboradores · status: {comp.status}"
                             + (f" · {len(rescisoes)} rescisão(ões) no mês, "
                                f"líquido R$ {resc_total:,.2f}" if rescisoes else ""),
                             ["Mat.", "Nome", "Regime", "Centro de Custo", "Bruto",
                              "INSS", "A pagar", "Custo total", "Situação"],
                             rows + [tot], [14, 56, 22, 38, 22, 20, 24, 26, 34],
                             _quem(request), f"folha_{comp.ano}_{comp.mes:02d}.pdf",
                             aligns_dir={4, 5, 6, 7}, linha_total=True, paisagem=True)

    headers = ["Mat.", "Nome", "Regime", "Centro de Custo", "Situação", "Sal. Bruto",
               "Faltas (d)", "Faltas (h)", "Desc. Faltas", "DSR perdido", "Desc. INSS",
               "Desc. IRRF", "Desc. VT", "VT", "VA", "Saldo Livre", "Prêmios", "Acerto",
               "Férias", "1/3 Férias", "Sal. Família", "Total a Pagar", "13º", "Férias prov.",
               "1/3 prov.", "FGTS", "Multa FGTS", "Recesso", "Patronal", "Custo Total"]
    rows = []
    for it in comp.itens.all():
        rows.append([it.matricula, it.nome, it.regime, it.centro_custo_nome, _marca(it),
                     it.salario_bruto, it.faltas_dias, it.faltas_horas, it.desc_faltas,
                     getattr(it, "desc_dsr", 0) or 0, it.desc_inss,
                     getattr(it, "desc_irrf", 0) or 0, it.desc_vt,
                     it.vt_com_faltas, it.va_com_faltas,
                     it.saldo_livre, it.premiacoes, it.acerto_contabil,
                     it.ferias_valor, it.ferias_terco, it.salario_familia, it.total_pagar,
                     it.decimo_mensal, it.ferias_mensal, it.terco_ferias_mensal,
                     it.fgts_mensal, it.multa_fgts_mensal, it.recesso_mensal,
                     it.inss_patronal, it.custo_total])
    wb, ws = _wb_timbrado("Folha Analítica",
                          f"Competência {rotulo} · {len(rows)} colaboradores · status: {comp.status}"
                          + (f" · {len(rescisoes)} rescisão(ões), líquido R$ {resc_total:,.2f}"
                             if rescisoes else ""),
                          _quem(request))
    _tabela(ws, 5, headers, rows,
            larguras=[8, 32, 11, 24, 26] + [12] * 25,
            money_cols=set(range(6, 31)) - {7, 8})

    # aba propria com as verbas de cada rescisao, linha a linha
    if rescisoes:
        wsr = wb.create_sheet("Rescisões")
        linhas_r = []
        for r in rescisoes:
            linhas_r.append([r.colaborador.matricula, r.colaborador.nome,
                             r.get_tipo_display(), str(r.data_desligamento),
                             r.aviso_dias, r.proventos, r.total_descontos, r.liquido,
                             (r.motivo or "")[:120]])
        linhas_r.append(["", f"TOTAL ({len(rescisoes)})", "", "", "",
                         round(sum(r.proventos for r in rescisoes), 2),
                         round(sum(r.total_descontos for r in rescisoes), 2),
                         resc_total, ""])
        _tabela(wsr, 1, ["Mat.", "Nome", "Tipo", "Desligamento", "Aviso (d)",
                         "Proventos (R$)", "Descontos (R$)", "Líquido (R$)", "Motivo"],
                linhas_r, larguras=[8, 32, 22, 14, 10, 15, 15, 15, 46],
                money_cols={6, 7, 8})

        # detalhe verba a verba, com a memoria de calculo congelada
        wsd = wb.create_sheet("Verbas rescisórias")
        det = []
        for r in rescisoes:
            for v in (r.verbas or []):
                det.append([r.colaborador.matricula, r.colaborador.nome, "Provento",
                            v.get("descricao", ""), v.get("valor", 0), v.get("memoria", "")])
            for d in (r.descontos or []):
                det.append([r.colaborador.matricula, r.colaborador.nome, "Desconto",
                            d.get("descricao", ""), d.get("valor", 0), d.get("memoria", "")])
        _tabela(wsd, 1, ["Mat.", "Nome", "Natureza", "Verba", "Valor (R$)", "Memória de cálculo"],
                det, larguras=[8, 32, 12, 34, 14, 70], money_cols={5})
    return _resposta_excel(wb, f"folha_{comp.ano}_{comp.mes:02d}.xlsx")


def _pdf_rateio(comp, rows, tot, usuario: str, rotulo: str) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)
    from reportlab.lib.styles import ParagraphStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            title=f"Rateio por Centro de Custo — {rotulo}")
    azul = colors.HexColor("#1E7BFF")
    navy = colors.HexColor("#0A1940")
    story = []
    if os.path.exists(_LOGO):
        story.append(Image(_LOGO, width=44 * mm, height=15.6 * mm, hAlign="LEFT"))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("Rateio de Pessoal por Centro de Custo",
                           ParagraphStyle("t", fontSize=16, textColor=navy, fontName="Helvetica-Bold")))
    story.append(Spacer(1, 1.5 * mm))
    story.append(Paragraph(
        f"Competência {rotulo} · status: {comp.status} · gerado por {usuario} em "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("s", fontSize=8, textColor=colors.HexColor("#666666"))))
    story.append(Spacer(1, 6 * mm))

    brl = lambda v: f"R$ {v:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    data = [["Centro de Custo", "HC", "Folha", "Provisões", "Patronal", "Custo Total"]]
    for r in rows:
        data.append([r[0], str(r[1]), brl(r[2]), brl(r[3]), brl(r[4]), brl(r[5])])
    data.append([tot[0], str(tot[1]), brl(tot[2]), brl(tot[3]), brl(tot[4]), brl(tot[5])])

    t = Table(data, colWidths=[62 * mm, 12 * mm, 26 * mm, 26 * mm, 26 * mm, 27 * mm], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F6FF")]),
        ("BACKGROUND", (0, -1), (-1, -1), azul),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCD6EE")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph("MDR Advocacia · Painel Financeiro — powered by Duna.Tech",
                           ParagraphStyle("f", fontSize=7, textColor=colors.HexColor("#999999"))))
    doc.build(story)
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="rateio_cc_{comp.ano}_{comp.mes:02d}.pdf"'
    resp["Access-Control-Expose-Headers"] = "Content-Disposition"
    return resp


def _pdf_generico(titulo: str, subtitulo: str, headers: list, rows: list, larguras: list,
                  usuario: str, nome_arquivo: str, aligns_dir: set = None,
                  linha_total: bool = False, paisagem: bool = False) -> HttpResponse:
    """PDF timbrado genérico (usado por quadro, folha, catálogos, simulação…)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4) if paisagem else A4,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm, title=titulo)
    navy = colors.HexColor("#0A1940")
    azul = colors.HexColor("#1E7BFF")
    story = []
    if os.path.exists(_LOGO):
        story.append(Image(_LOGO, width=40 * mm, height=14.2 * mm, hAlign="LEFT"))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(titulo, ParagraphStyle("t", fontSize=15, textColor=navy,
                                                  fontName="Helvetica-Bold")))
    story.append(Paragraph(
        f"{subtitulo} · gerado por {usuario} em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("s", fontSize=7.5, textColor=colors.HexColor("#666666"))))
    story.append(Spacer(1, 5 * mm))

    data = [headers] + [[str(c) for c in r] for r in rows]
    t = Table(data, colWidths=[w * mm for w in larguras], repeatRows=1)
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1 if not linha_total else -2),
         [colors.white, colors.HexColor("#F2F6FF")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCD6EE")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for c in (aligns_dir or set()):
        estilo.append(("ALIGN", (c, 0), (c, -1), "RIGHT"))
    if linha_total:
        estilo += [("BACKGROUND", (0, -1), (-1, -1), azul),
                   ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
                   ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")]
    t.setStyle(TableStyle(estilo))
    story.append(t)
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("MDR Advocacia · Painel Financeiro — powered by Duna.Tech",
                           ParagraphStyle("f", fontSize=7, textColor=colors.HexColor("#999999"))))
    doc.build(story)
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    resp["Access-Control-Expose-Headers"] = "Content-Disposition"
    return resp


def _brl(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    except (TypeError, ValueError):
        return str(v)


@api_view(["GET"])
@permission_classes(_PERM)
def dp_relatorio_catalogos(request):
    """Cargos + Centros de Custo (Excel com 2 abas, ou PDF)."""
    from .models import DpCargo
    cargos = list(DpCargo.objects.all())
    ccs = list(DpCentroCusto.objects.all())
    usuario = _quem(request)

    if request.query_params.get("formato") == "pdf":
        rows = [[c.area, c.nome, _brl(c.salario_base), c.dias_mes, c.carga_horaria_mes] for c in cargos]
        return _pdf_generico("Plano de Cargos", f"{len(rows)} cargos",
                             ["Área", "Cargo", "Salário base", "Dias/mês", "Carga/mês"],
                             rows, [20, 80, 30, 20, 25], usuario, "plano_cargos.pdf",
                             aligns_dir={2, 3, 4})

    wb, ws = _wb_timbrado("Plano de Cargos", f"{len(cargos)} cargos cadastrados", usuario)
    _tabela(ws, 5, ["Área", "Cargo", "Salário base (R$)", "Dias/mês", "Carga horária/mês", "Ativo"],
            [[c.area, c.nome, c.salario_base, c.dias_mes, c.carga_horaria_mes,
              "Sim" if c.ativo else "Não"] for c in cargos],
            larguras=[10, 40, 18, 11, 18, 8], money_cols={3})
    ws2 = wb.create_sheet("Centros de Custo")
    _tabela(ws2, 1, ["Código", "Centro de Custo", "Colaboradores ativos"],
            [[c.codigo, c.nome, c.colaboradores.filter(status="ativo").count()] for c in ccs],
            larguras=[10, 40, 20])
    return _resposta_excel(wb, "cargos_e_centros_custo.xlsx")


@api_view(["GET"])
@permission_classes(_PERM)
def dp_relatorio_auditoria(request):
    """Trilha de auditoria em Excel (últimos N registros)."""
    from .models import DpAuditLog
    try:
        limit = min(int(request.query_params.get("limit", 2000)), 20000)
    except ValueError:
        limit = 2000
    logs = DpAuditLog.objects.all()[:limit]
    rows = []
    for a in logs:
        h = humanizar(a)
        detalhe = "; ".join(
            (f"{m['campo']}: {m['de']} -> {m['para']}" if m["de"] else f"{m['campo']}: {m['para']}")
            for m in h["mudancas"]) or "—"
        rows.append([h["quando_br"], h["usuario"], h["titulo"], detalhe])
    wb, ws = _wb_timbrado("Trilha de Auditoria — Controle de Pessoal",
                          f"{len(rows)} registro(s) mais recentes", _quem(request))
    _tabela(ws, 5, ["Quando", "Quem", "O que aconteceu", "Detalhes da alteração"],
            rows, larguras=[20, 28, 62, 78])
    return _resposta_excel(wb, "auditoria_dp.xlsx")


@api_view(["GET"])
@permission_classes(_PERM)
def dp_relatorio_dashboard(request):
    """Dashboard do DP em Excel: KPIs + série de custo + movimentação."""
    from .dp_relatorios import dp_dashboard as _dash  # reusa o cálculo
    dados = _dash(request._request if hasattr(request, "_request") else request).data
    usuario = _quem(request)
    wb, ws = _wb_timbrado("Dashboard — Controle de Pessoal",
                          f"Headcount ativo: {dados['headcount']}", usuario)
    _tabela(ws, 5, ["Indicador", "Valor"], [
        ["Headcount ativo", dados["headcount"]],
        ["Admissões no mês", dados["admissoes_mes"]],
        ["Desligamentos no mês", dados["desligamentos_mes"]],
        ["Turnover do mês (%)", dados["turnover_mes"]],
    ] + [[f"Headcount — {REG_LABEL.get(k, k)}", v] for k, v in dados["por_regime"].items()],
        larguras=[34, 16])
    ws2 = wb.create_sheet("Evolução mensal")
    _tabela(ws2, 1, ["Competência", "Situação", "Colaboradores", "Admissões", "Desligamentos",
                     "Rotatividade (%)", "Custo total (R$)", "Custo CLT (R$)", "Provisões (R$)",
                     "INSS patronal (R$)", "FGTS do mês (R$)", "FGTS acumulado (R$)",
                     "Multa FGTS (R$)", "Multa acumulada (R$)"],
            [[l["mes"], l["status"], l["headcount"], l.get("admissoes", 0), l.get("desligamentos", 0),
              l.get("turnover", 0), l["custo_total"], l.get("custo_clt", 0), l["provisoes"],
              l["patronal"], l.get("fgts", 0), l.get("fgts_acumulado", 0),
              l.get("multa_fgts", 0), l.get("multa_fgts_acumulada", 0)]
             for l in dados["serie_custo"]],
            larguras=[14, 12, 13, 11, 13, 15, 17, 16, 15, 17, 15, 18, 15, 18],
            money_cols={7, 8, 9, 10, 11, 12, 13, 14})
    if dados.get("custo_por_cargo"):
        wsc = wb.create_sheet("Custo por cargo")
        _tabela(wsc, 1, ["Cargo", "Tipo de contrato", "Pessoas", "Salário médio (R$)",
                         "Custo total (R$)", "Custo médio por pessoa (R$)"],
                [[c["cargo"], REG_LABEL.get(c["regime"], c["regime"]), c["quantidade"],
                  c["salario_medio"], c["custo"], c["custo_medio"]]
                 for c in dados["custo_por_cargo"]],
                larguras=[34, 16, 10, 18, 18, 24], money_cols={4, 5, 6})
    ws3 = wb.create_sheet("Movimentação")
    _tabela(ws3, 1, ["Mês", "Admissões", "Desligamentos"],
            [[l["mes"], l["admissoes"], l["desligamentos"]] for l in dados["serie_mov"]],
            larguras=[14, 14, 16])
    return _resposta_excel(wb, "dashboard_dp.xlsx")


REG_LABEL = {"estagiario": "Estagiário (TCE)", "clt": "CLT",
             "associado": "Associado", "pj": "PJ"}


@api_view(["POST"])
@permission_classes(_PERM)
def dp_relatorio_simulacao(request):
    """Recebe o RESULTADO da simulação (do front) e devolve Excel ou PDF timbrado."""
    d = request.data or {}
    formato = request.query_params.get("formato", "excel")
    nome = d.get("nome") or "Cenário"
    usuario = _quem(request)
    atual, cen, delta = d.get("atual", {}), d.get("cenario", {}), d.get("delta", {})
    resumo = [
        ["Headcount", atual.get("headcount", 0), cen.get("headcount", 0), delta.get("headcount", 0)],
        ["Folha (a pagar)", atual.get("folha", 0), cen.get("folha", 0), delta.get("folha", 0)],
        ["Provisões", atual.get("provisoes", 0), cen.get("provisoes", 0), delta.get("provisoes", 0)],
        ["INSS patronal", atual.get("patronal", 0), cen.get("patronal", 0), delta.get("patronal", 0)],
        ["CUSTO TOTAL", atual.get("custo_total", 0), cen.get("custo_total", 0), delta.get("custo_total", 0)],
    ]
    sub = (f"Impacto mensal: {_brl(d.get('impacto_mensal', 0))} · "
           f"anual ({d.get('meses', 12)} meses): {_brl(d.get('impacto_anual', 0))}")

    if formato == "pdf":
        rows = [[r[0], _brl(r[1]) if r[0] != "Headcount" else r[1],
                 _brl(r[2]) if r[0] != "Headcount" else r[2],
                 _brl(r[3]) if r[0] != "Headcount" else r[3]] for r in resumo]
        return _pdf_generico(f"Simulação — {nome}", sub,
                             ["Indicador", "Cenário atual", "Cenário simulado", "Variação"],
                             rows, [55, 40, 40, 40], usuario, "simulacao.pdf",
                             aligns_dir={1, 2, 3}, linha_total=True)

    wb, ws = _wb_timbrado(f"Simulação — {nome}", sub, usuario)
    _tabela(ws, 5, ["Indicador", "Cenário atual", "Cenário simulado", "Variação"],
            resumo, larguras=[28, 20, 20, 18], money_cols={2, 3, 4})
    novos = d.get("detalhe_novos") or []
    if novos:
        ws2 = wb.create_sheet("Contratações simuladas")
        _tabela(ws2, 1, ["Vaga", "Regime", "Centro de Custo", "Salário bruto",
                         "A pagar", "Provisões", "Patronal", "Custo total"],
                [[n["nome"], REG_LABEL.get(n["regime"], n["regime"]), n["cc"], n["salario_bruto"],
                  n["total_pagar"], n["provisoes"], n["patronal"], n["custo_total"]] for n in novos],
                larguras=[26, 14, 24, 15, 14, 14, 14, 16], money_cols={4, 5, 6, 7, 8})
    porcc = d.get("por_centro_custo") or []
    if porcc:
        ws3 = wb.create_sheet("Impacto por CC")
        _tabela(ws3, 1, ["Centro de Custo", "Novas vagas", "Custo mensal"],
                [[c["centro_custo"], c["headcount"], c["custo_total"]] for c in porcc],
                larguras=[32, 14, 18], money_cols={3})
    return _resposta_excel(wb, "simulacao.xlsx")


@api_view(["GET"])
@permission_classes(_PERM)
def dp_relatorio_projecao(request):
    """Projeção de gastos + aprovisionamento em Excel ou PDF."""
    from .dp_simulacao import dp_projecao as _proj
    dados = _proj(request._request if hasattr(request, "_request") else request).data
    usuario = _quem(request)
    p = dados["premissas"]
    sub = (f"{p['meses']} meses · reajuste {p['reajuste']:.1%} no mês {p['mes_reajuste']} · "
           f"crescimento {p['crescimento']:.1%} a.m.")
    linhas = dados["linhas"]

    if request.query_params.get("formato") == "pdf":
        rows = [[l["mes"], l["headcount"], _brl(l["folha"]), _brl(l["provisoes"]),
                 _brl(l["patronal"]), _brl(l["custo_total"]), _brl(l["provisionado_acumulado"])]
                for l in linhas]
        return _pdf_generico("Projeção de Gastos com Pessoal", sub,
                             ["Mês", "HC", "Folha", "Provisões", "Patronal", "Custo total",
                              "Provisionado acum."],
                             rows, [20, 14, 30, 30, 28, 32, 36], usuario, "projecao.pdf",
                             aligns_dir={1, 2, 3, 4, 5, 6}, paisagem=True)

    wb, ws = _wb_timbrado("Projeção de Gastos com Pessoal", sub, usuario)
    _tabela(ws, 5, ["Mês", "Headcount", "Folha (R$)", "Provisões (R$)", "Patronal (R$)",
                    "Custo total (R$)", "Provisionado acumulado (R$)"],
            [[l["mes"], l["headcount"], l["folha"], l["provisoes"], l["patronal"],
              l["custo_total"], l["provisionado_acumulado"]] for l in linhas],
            larguras=[12, 12, 16, 16, 16, 18, 24], money_cols={3, 4, 5, 6, 7})
    ap = dados["aprovisionamento"]
    ws2 = wb.create_sheet("Aprovisionamento")
    _tabela(ws2, 1, ["Verba provisionada", "Valor acumulado (R$)"],
            [["13º salário", ap["decimo"]], ["Férias", ap["ferias"]], ["1/3 de férias", ap["terco"]],
             ["FGTS", ap["fgts"]], ["Multa FGTS (40%)", ap["multa_fgts"]],
             ["Recesso (estagiários)", ap["recesso"]], ["TOTAL", ap["total"]]],
            larguras=[30, 24], money_cols={2})
    return _resposta_excel(wb, "projecao_gastos.xlsx")


def _pacote(c) -> float:
    """Quanto a pessoa recebe no pacote CHEIO: salário + saldo livre + VT + VA.

    É valor BRUTO — antes de INSS e do desconto do vale-transporte. Não é o
    líquido da folha; existe pra responder "quanto essa pessoa ganha" sem
    obrigar a somar quatro colunas na mão.
    """
    return round((c.salario_bruto or 0) + (c.saldo_livre or 0)
                 + (c.vt or 0) + (c.va or 0), 2)


@api_view(["GET"])
@permission_classes(_PERM)
def dp_relatorio_quadro(request):
    """Excel/PDF do Quadro de Pessoal, com OS MESMOS filtros da tela.

    Os decorators aqui não são enfeite. Sem @api_view, `request` chega como
    WSGIRequest e `request.query_params` não existe — era o 500 em toda
    exportação. E sem @permission_classes o endpoint ficava SEM autenticação
    nenhuma: só não vazava o quadro (nome, CPF, salário) porque estourava antes
    de responder. Confirmado em produção: anônimo entrava na view; os endpoints
    irmãos devolvem 401.
    """
    p = request.query_params
    status_f = p.get("status", "")
    qs = filtrar_quadro(
        filtrar_colaboradores(
            DpColaborador.objects.select_related("centro_custo", "cargo").all(),
            request.user),
        p)
    # o subtítulo tem que dizer o recorte, senão duas planilhas diferentes
    # saem com o mesmo cabeçalho e ninguém sabe qual é qual
    recorte = " · ".join(
        f"{rot}: {p[k]}" for k, rot in
        (("status", "situação"), ("regime", "regime"), ("busca", "busca"),
         ("unidade", "unidade"))
        if p.get(k)) or "todos"
    if p.get("formato") == "pdf":
        rows = [[c.matricula, c.nome[:32], REG_LABEL.get(c.regime, c.regime), c.status,
                 c.centro_custo.nome[:24] if c.centro_custo_id else "",
                 (c.cargo.nome[:24] if c.cargo_id else ""), _brl(c.salario_bruto),
                 _brl(c.saldo_livre), _brl(_pacote(c))] for c in qs]
        return _pdf_generico("Quadro de Pessoal",
                             f"{len(rows)} colaborador(es) · {recorte}",
                             ["Mat.", "Nome", "Regime", "Status", "Centro de Custo", "Cargo",
                              "Sal. bruto", "Saldo livre", "Total"],
                             rows, [15, 60, 24, 16, 46, 46, 26, 26, 27], _quem(request),
                             "quadro_pessoal.pdf", aligns_dir={6, 7, 8}, paisagem=True)

    headers = ["Mat.", "Nome", "CPF", "Regime", "Status", "Unidade", "Área",
               "Centro de Custo", "Supervisor", "Equipe", "Cargo", "Admissão",
               "Demissão", "Sal. Bruto", "Saldo Livre", "VT", "VA", "Total"]
    rows = []
    for c in qs:
        rows.append([c.matricula, c.nome, c.cpf, c.get_regime_display(), c.status,
                     c.unidade, c.area, c.centro_custo.nome if c.centro_custo_id else "",
                     c.supervisor.nome if c.supervisor_id else "", c.equipe,
                     c.cargo.nome if c.cargo_id else "",
                     str(c.data_admissao or ""), str(c.data_demissao or ""),
                     c.salario_bruto, c.saldo_livre, c.vt, c.va, _pacote(c)])
    wb, ws = _wb_timbrado("Quadro de Pessoal",
                          f"{len(rows)} colaborador(es) · {recorte}",
                          _quem(request))
    _tabela(ws, 5, headers, rows,
            larguras=[8, 32, 13, 13, 9, 13, 7, 24, 16, 18, 24, 11, 11, 12, 11, 9, 9, 12],
            money_cols={14, 15, 16, 17, 18})
    return _resposta_excel(wb, "quadro_pessoal.xlsx")


def _pdf_termo_rescisao(r, usuario: str) -> HttpResponse:
    """Termo de rescisão timbrado: identificação, verbas, descontos e líquido."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    c = r.colaborador
    navy = colors.HexColor("#0A1940")
    azul = colors.HexColor("#1E7BFF")
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            title=f"Termo de rescisão — {c.nome}")
    st_lbl = ParagraphStyle("l", fontSize=8, textColor=colors.HexColor("#666666"))
    story = []
    if os.path.exists(_LOGO):
        story.append(Image(_LOGO, width=42 * mm, height=14.9 * mm, hAlign="LEFT"))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("Termo de Rescisão do Contrato de Trabalho",
                           ParagraphStyle("t", fontSize=15, textColor=navy, fontName="Helvetica-Bold")))
    story.append(Paragraph(
        f"Gerado por {usuario} em {datetime.now().strftime('%d/%m/%Y %H:%M')} · MDR Advocacia",
        st_lbl))
    story.append(Spacer(1, 5 * mm))

    def dbr(d):
        return d.strftime("%d/%m/%Y") if d else "—"

    ident = [
        ["Colaborador", c.nome, "Matrícula", str(c.matricula)],
        ["CPF", c.cpf or "—", "Tipo de contrato", dict(
            [("estagiario", "Estagiário (TCE)"), ("clt", "CLT"),
             ("associado", "Associado"), ("pj", "PJ")]).get(c.regime, c.regime)],
        ["Cargo", (c.cargo.nome if c.cargo_id else "—"), "Centro de custo",
         (c.centro_custo.nome if c.centro_custo_id else "—")],
        ["Admissão", dbr(c.data_admissao or c.data_entrada), "Desligamento", dbr(r.data_desligamento)],
        ["Motivo", r.get_tipo_display(), "Aviso prévio",
         f"{r.aviso_dias} dias" if r.aviso_dias else "—"],
    ]
    t = Table(ident, colWidths=[26 * mm, 62 * mm, 30 * mm, 60 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#666666")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#666666")),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, colors.HexColor("#EEEEEE")),
    ]))
    story.append(t)
    story.append(Spacer(1, 6 * mm))

    def bloco(titulo, linhas, cor, total_label, total):
        story.append(Paragraph(titulo, ParagraphStyle("b", fontSize=10, textColor=cor,
                                                      fontName="Helvetica-Bold")))
        story.append(Spacer(1, 1.5 * mm))
        data = [["Verba", "Como foi calculado", "Valor"]]
        for v in linhas:
            data.append([v["descricao"], v.get("memoria", ""), _brl(v["valor"])])
        data.append([total_label, "", _brl(total)])
        tb = Table(data, colWidths=[52 * mm, 90 * mm, 36 * mm], repeatRows=1)
        tb.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), navy),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("TEXTCOLOR", (1, 1), (1, -2), colors.HexColor("#777777")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7F9FF")]),
            ("BACKGROUND", (0, -1), (-1, -1), cor),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCD6EE")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(tb)
        story.append(Spacer(1, 5 * mm))

    bloco("Verbas devidas (proventos)", r.verbas or [], azul, "Total de proventos", r.proventos)
    if r.descontos:
        bloco("Descontos", r.descontos, colors.HexColor("#C0392B"), "Total de descontos", r.total_descontos)

    liq = Table([["VALOR LÍQUIDO A RECEBER", _brl(r.liquido)]], colWidths=[130 * mm, 48 * mm])
    liq.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0A1940")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(liq)
    if r.motivo:
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(f"<b>Observações:</b> {r.motivo}", ParagraphStyle("o", fontSize=8)))
    story.append(Spacer(1, 14 * mm))
    ass = Table([["_" * 42, "_" * 42], ["MDR Advocacia", c.nome]],
                colWidths=[85 * mm, 85 * mm])
    ass.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 8),
                             ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                             ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#666666"))]))
    story.append(ass)
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        "Documento gerado pelo Painel Financeiro — confira os valores antes da homologação.",
        ParagraphStyle("f", fontSize=7, textColor=colors.HexColor("#999999"))))
    doc.build(story)
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="rescisao_{c.matricula}.pdf"'
    resp["Access-Control-Expose-Headers"] = "Content-Disposition"
    return resp


# ────────────────── FICHA FINANCEIRA DO COLABORADOR (PDF) ──────────────────

@api_view(["GET"])
@permission_classes(_PERM)
def dp_ficha_financeira(request, pk):
    """Histórico completo de recebimentos de UMA pessoa, mês a mês, em PDF
    timbrado: o que entrou, o que foi descontado, benefícios, custo pro
    escritório, médias e — se houver — a rescisão."""
    from .models import DpColaborador, DpFolhaItem, DpRescisao

    colab = DpColaborador.objects.filter(pk=pk).select_related(
        "centro_custo", "cargo", "supervisor", "coordenador").first()
    if not colab:
        return Response({"detail": "Colaborador não encontrado."}, status=404)
    if not filtrar_colaboradores(DpColaborador.objects.filter(pk=pk), request.user).exists():
        return Response({"detail": "Fora do seu escopo de acesso."}, status=403)

    itens = (DpFolhaItem.objects.filter(colaborador=colab)
             .select_related("competencia")
             .order_by("competencia__ano", "competencia__mes"))
    rescisao = DpRescisao.objects.filter(colaborador=colab).order_by("-data_desligamento").first()
    return _pdf_ficha_financeira(colab, list(itens), rescisao, _quem(request))


def _pdf_ficha_financeira(colab, itens, rescisao, usuario: str) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, KeepTogether, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    navy = colors.HexColor("#0A1940")
    azul = colors.HexColor("#1E7BFF")
    cinza = colors.HexColor("#666666")
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=11 * mm,
                            bottomMargin=11 * mm, leftMargin=11 * mm, rightMargin=11 * mm,
                            title=f"Ficha financeira — {colab.nome}")
    st_tit = ParagraphStyle("t", fontSize=15, textColor=navy, fontName="Helvetica-Bold")
    st_sub = ParagraphStyle("s", fontSize=7.5, textColor=cinza)
    st_sec = ParagraphStyle("sec", fontSize=9.5, textColor=navy, fontName="Helvetica-Bold",
                            spaceBefore=4, spaceAfter=2)
    st_min = ParagraphStyle("m", fontSize=7, textColor=cinza)

    story = []
    if os.path.exists(_LOGO):
        story.append(Image(_LOGO, width=40 * mm, height=14.2 * mm, hAlign="LEFT"))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph("Ficha financeira do colaborador", st_tit))
    story.append(Paragraph(
        f"Documento interno · gerado por {usuario} em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        st_sub))
    story.append(Spacer(1, 4 * mm))

    # ── bloco cadastral ──
    def _d(v):
        return v.strftime("%d/%m/%Y") if v else "—"

    ident = [
        ["Nome", colab.nome, "Matrícula", str(colab.matricula)],
        ["Contrato", colab.get_regime_display(), "Situação",
         "Ativo" if colab.status == "ativo" else f"Desligado em {_d(colab.data_demissao)}"],
        ["Cargo", colab.cargo.nome if colab.cargo_id else "—",
         "Centro de custo", colab.centro_custo.nome if colab.centro_custo_id else "—"],
        ["Unidade", colab.unidade or "—", "Admissão", _d(colab.data_admissao or colab.data_entrada)],
        ["Supervisor", colab.supervisor.nome if colab.supervisor_id else "—",
         "Coordenador", colab.coordenador.nome if colab.coordenador_id else "—"],
    ]
    t_id = Table(ident, colWidths=[26 * mm, 92 * mm, 30 * mm, 87 * mm])
    t_id.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (0, 0), (0, -1), cinza), ("TEXTCOLOR", (2, 0), (2, -1), cinza),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F6FF")),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCD6EE")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t_id)
    story.append(Spacer(1, 5 * mm))

    # ── histórico mês a mês ──
    story.append(Paragraph("Histórico de recebimentos", st_sec))
    headers = ["Competência", "Salário", "Faltas", "Desc. faltas", "INSS", "Desc. VT",
               "Vale-transp.", "Vale-alim.", "Saldo livre", "Prêmios", "Sal. família",
               "Acertos", "Líquido a pagar", "Custo total"]
    linhas, tot = [], {k: 0.0 for k in
                       ("sal", "descf", "inss", "descvt", "vt", "va", "saldo", "prem",
                        "salfam", "acerto", "pagar", "custo")}
    for it in itens:
        comp = it.competencia
        faltas = []
        if it.faltas_dias:
            faltas.append(f"{it.faltas_dias:g}d")
        if it.faltas_horas:
            faltas.append(f"{it.faltas_horas:g}h")
        linhas.append([
            f"{comp.mes:02d}/{comp.ano}", _brl(it.salario_bruto), " ".join(faltas) or "—",
            _brl(it.desc_faltas), _brl(it.desc_inss), _brl(it.desc_vt),
            _brl(getattr(it, "vt_com_faltas", None) or it.vt),
            _brl(getattr(it, "va_com_faltas", None) or it.va),
            _brl(it.saldo_livre), _brl(it.premiacoes),
            _brl(getattr(it, "salario_familia", 0) or 0), _brl(it.acerto_contabil),
            _brl(it.total_pagar), _brl(it.custo_total),
        ])
        tot["sal"] += it.salario_bruto or 0
        tot["descf"] += it.desc_faltas or 0
        tot["inss"] += it.desc_inss or 0
        tot["descvt"] += it.desc_vt or 0
        tot["vt"] += getattr(it, "vt_com_faltas", None) or it.vt or 0
        tot["va"] += getattr(it, "va_com_faltas", None) or it.va or 0
        tot["saldo"] += it.saldo_livre or 0
        tot["prem"] += it.premiacoes or 0
        tot["salfam"] += getattr(it, "salario_familia", 0) or 0
        tot["acerto"] += it.acerto_contabil or 0
        tot["pagar"] += it.total_pagar or 0
        tot["custo"] += it.custo_total or 0

    if linhas:
        linhas.append(["TOTAL", _brl(tot["sal"]), "", _brl(tot["descf"]), _brl(tot["inss"]),
                       _brl(tot["descvt"]), _brl(tot["vt"]), _brl(tot["va"]), _brl(tot["saldo"]),
                       _brl(tot["prem"]), _brl(tot["salfam"]), _brl(tot["acerto"]),
                       _brl(tot["pagar"]), _brl(tot["custo"])])
        # ATENÇÃO: as larguras são em MILÍMETROS e precisam do `* mm`. Sem ele o
        # reportlab lê como pontos, a tabela encolhe pra ~1/3 da página e as
        # colunas se sobrepõem (foi exatamente o bug reportado na ficha).
        # Soma = 263mm; a paisagem A4 útil é 275mm (297 − 11 − 11 de margem).
        larguras = [19, 20, 12, 19, 19, 18, 19, 19, 19, 18, 19, 18, 23, 21]
        t = Table([headers] + linhas, colWidths=[w * mm for w in larguras], repeatRows=1)
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), navy),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F6FF")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCD6EE")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("BACKGROUND", (0, -1), (-1, -1), azul),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]
        # destaca os meses com premiação (o que o operador procura na hora)
        for i, it in enumerate(itens, start=1):
            if it.premiacoes:
                estilo.append(("BACKGROUND", (9, i), (9, i), colors.HexColor("#DCFCE7")))
                estilo.append(("TEXTCOLOR", (9, i), (9, i), colors.HexColor("#15803D")))
        t.setStyle(TableStyle(estilo))
        story.append(t)

        # ── resumo ──
        n = len(itens)
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("Resumo do período", st_sec))
        meses_prem = sum(1 for i in itens if i.premiacoes)
        resumo = [
            ["Meses apurados", str(n),
             "Média mensal recebida", _brl(tot["pagar"] / n if n else 0)],
            ["Total recebido no período", _brl(tot["pagar"]),
             "Custo médio pro escritório", _brl(tot["custo"] / n if n else 0)],
            ["Total de premiações", _brl(tot["prem"]),
             "Meses com premiação", f"{meses_prem} de {n}"],
            ["Total descontado (faltas + INSS + VT)",
             _brl(tot["descf"] + tot["inss"] + tot["descvt"]),
             "Benefícios recebidos (VT + VA)", _brl(tot["vt"] + tot["va"])],
        ]
        t_r = Table(resumo, colWidths=[62 * mm, 52 * mm, 62 * mm, 59 * mm])
        t_r.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TEXTCOLOR", (0, 0), (0, -1), cinza), ("TEXTCOLOR", (2, 0), (2, -1), cinza),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
            ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"), ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F6FF")),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCD6EE")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDE5F5")),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t_r)
    else:
        story.append(Paragraph(
            "Esta pessoa ainda não aparece em nenhuma folha calculada.", st_min))

    # ── rescisão, quando houver ──
    if rescisao:
        story.append(Spacer(1, 5 * mm))
        bloco = [Paragraph("Rescisão", st_sec)]
        verbas = [[v.get("descricao", ""), _brl(v.get("valor", 0))]
                  for v in (rescisao.verbas or [])]
        descontos = [[d.get("descricao", ""), f"- {_brl(d.get('valor', 0))}"]
                     for d in (rescisao.descontos or [])]
        dados = ([["Desligamento", _d(rescisao.data_desligamento)],
                  ["Motivo", rescisao.get_tipo_display()]]
                 + verbas + descontos
                 + [["LÍQUIDO DA RESCISÃO", _brl(rescisao.liquido)]])
        t_v = Table(dados, colWidths=[150 * mm, 40 * mm])
        t_v.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCD6EE")),
            ("BACKGROUND", (0, -1), (-1, -1), azul),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        bloco.append(t_v)
        story.append(KeepTogether(bloco))

    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(
        "Documento gerado pelo Controle de Pessoal — uso interno. Os valores refletem as "
        "competências calculadas no sistema na data da emissão.", st_min))
    story.append(Paragraph("MDR Advocacia · Painel Financeiro — powered by Duna.Tech", st_min))
    doc.build(story)

    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    nome = f"ficha-financeira-{colab.matricula}.pdf"
    resp["Content-Disposition"] = f'attachment; filename="{nome}"'
    resp["Access-Control-Expose-Headers"] = "Content-Disposition"
    return resp
