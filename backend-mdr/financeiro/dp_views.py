# Módulo Controle de Pessoal (DP) — F1: cadastro + importador + auditoria.
# RBAC: leitura exige nível "ver" no módulo `pessoal`; escrita exige "editar".
# TODA escrita gera DpAuditLog (imutável) e, quando cadastral, DpEvento.
from datetime import date, datetime

from django.conf import settings
from django.db import transaction
from django.db.models import Count, Max, Q
from django.http import FileResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from .models import (
    DP_MATRICULA_BASE, DpAuditLog, DpCargo, DpCentroCusto, DpColaborador, DpDependente,
    DpAfastamento, DpDocumento, DpEvento, DpLideranca, DpTransferenciaContrato,
)
from .models import REGRAS_AFASTAMENTO
from .dp_audit import humanizar
from .dp_escopo import filtrar_colaboradores
from .serializers import (
    DpCargoSerializer, DpCentroCustoSerializer, DpColaboradorSerializer, DpLiderancaSerializer,
)
from .views import modulo_permission

_PERM = [modulo_permission(read_any=["pessoal"], write="pessoal")]


def _quem(request) -> str:
    u = request.user
    return (u.email or u.username or str(u.pk)) if u and u.is_authenticated else "?"


def audit(request, acao: str, entidade: str, entidade_id, antes=None, depois=None,
          colaborador=None):
    """Registra na trilha. `colaborador` liga o registro à PESSOA afetada — é o
    que permite pesquisar o histórico completo de alguém (cadastro, folha,
    ajuste pontual, rescisão). Quando a entidade já é a ficha, deduz sozinho."""
    if colaborador is None and entidade == "dp_colaborador" and entidade_id:
        colaborador = DpColaborador.objects.filter(pk=entidade_id).first()
    DpAuditLog.objects.create(
        usuario=_quem(request), acao=acao, entidade=entidade,
        entidade_id=str(entidade_id or ""), antes=antes, depois=depois,
        colaborador=colaborador,
        colaborador_nome=(colaborador.nome if colaborador else ""),
    )


def _data_ou_none(v):
    """Aceita 'AAAA-MM-DD' (ou vazio) e devolve date/None sem estourar."""
    if not v:
        return None
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _dependente_json(d: DpDependente, colab=None) -> dict:
    """Além dos campos, devolve o que a TELA precisa decidir sem recalcular:
    se hoje ele gera cota e se a comprovação está pendente."""
    hoje = date.today()
    idade = (hoje.year - d.data_nascimento.year
             - ((hoje.month, hoje.day) < (d.data_nascimento.month, d.data_nascimento.day)))
    ano14, mes14 = d.faz_14_em()
    return {
        "id": str(d.id),
        "nome": d.nome,
        "data_nascimento": d.data_nascimento.isoformat(),
        "nascimento_br": f"{d.data_nascimento:%d/%m/%Y}",
        "idade": idade,
        "tipo": d.tipo, "tipo_label": d.get_tipo_display(),
        "cpf": d.cpf, "invalido": d.invalido, "ativo": d.ativo,
        "observacao": d.observacao,
        "vacinacao_valida_ate": (d.vacinacao_valida_ate.isoformat()
                                 if d.vacinacao_valida_ate else None),
        "frequencia_escolar_valida_ate": (d.frequencia_escolar_valida_ate.isoformat()
                                          if d.frequencia_escolar_valida_ate else None),
        # gera cota HOJE? (só idade/tipo — a renda do mês é testada na folha)
        "elegivel_hoje": d.elegivel_em(hoje.year, hoje.month),
        "cota_ate": (None if d.invalido else f"{mes14:02d}/{ano14}"),
        "comprovacao_pendente": d.comprovacao_pendente_em(hoje),
    }


def _afastamento_json(a: DpAfastamento) -> dict:
    r = a.regra
    hoje = date.today()
    em_curso = not a.data_retorno and a.data_inicio <= hoje
    return {
        "id": str(a.id), "tipo": a.tipo, "tipo_label": a.get_tipo_display(),
        "data_inicio": a.data_inicio.isoformat(),
        "inicio_br": f"{a.data_inicio:%d/%m/%Y}",
        "data_prevista_retorno": (a.data_prevista_retorno.isoformat()
                                  if a.data_prevista_retorno else None),
        "data_retorno": a.data_retorno.isoformat() if a.data_retorno else None,
        "retorno_br": f"{a.data_retorno:%d/%m/%Y}" if a.data_retorno else "",
        "estabilidade_ate": a.estabilidade_ate.isoformat() if a.estabilidade_ate else None,
        "estabilidade_br": f"{a.estabilidade_ate:%d/%m/%Y}" if a.estabilidade_ate else "",
        "em_estabilidade": bool(a.estabilidade_ate and a.estabilidade_ate >= hoje),
        "observacao": a.observacao,
        "em_curso": em_curso,
        # o que a regra faz na folha, em português, pra tela poder explicar
        "regra": {
            "dias_empresa": r["dias_empresa"],
            "fgts": r["fgts"],
            "corta_va": r["corta_va"],
            "compensa_na_guia": r["compensa_na_guia"],
        },
    }


def _doc_json(d: DpDocumento) -> dict:
    return {
        "id": str(d.id), "tipo": d.tipo, "tipo_label": d.get_tipo_display(),
        "nome_original": d.nome_original, "tamanho": d.tamanho,
        "descricao": d.descricao, "enviado_por": d.enviado_por,
        "created_at": d.created_at.isoformat(),
        "quando_br": d.created_at.strftime("%d/%m/%Y às %H:%M"),
    }


def _snap(obj: DpColaborador) -> dict:
    return {
        "matricula": obj.matricula, "nome": obj.nome, "cpf": obj.cpf,
        "unidade": obj.unidade, "area": obj.area,
        "centro_custo": obj.centro_custo.nome if obj.centro_custo_id else None,
        "equipe_estrutura": obj.equipe_ref.nome if obj.equipe_ref_id else None,
        "equipe_id": str(obj.equipe_ref_id) if obj.equipe_ref_id else None,
        "supervisor": obj.supervisor.nome if obj.supervisor_id else None,
        "coordenador": obj.coordenador.nome if obj.coordenador_id else None,
        "equipe": obj.equipe,
        "cargo": obj.cargo.nome if obj.cargo_id else None,
        "regime": obj.regime, "status": obj.status,
        "data_admissao": str(obj.data_admissao or ""), "data_demissao": str(obj.data_demissao or ""),
        "salario_bruto": obj.salario_bruto, "saldo_livre": obj.saldo_livre,
        "vt": obj.vt, "va": obj.va, "opta_vt": obj.opta_vt,
        "aprendiz": obj.aprendiz,
    }


def proxima_matricula(regime: str) -> int:
    """Matrícula automática pela lógica da casa: 10xx/20xx/30xx/40xx por regime."""
    base = DP_MATRICULA_BASE.get(regime, 9000)
    atual = DpColaborador.objects.filter(
        matricula__gte=base, matricula__lt=base + 1000
    ).aggregate(m=Max("matricula"))["m"]
    return (atual + 1) if atual else base + 1


class DpCentroCustoViewSet(viewsets.ModelViewSet):
    queryset = DpCentroCusto.objects.all()
    serializer_class = DpCentroCustoSerializer
    permission_classes = _PERM

    @action(detail=False, methods=["get"])
    def arvore(self, request):
        """Centros de custo em árvore (núcleo → subnúcleos) com totais somados."""
        todos = list(DpCentroCusto.objects.select_related("pai").all())
        ativos = {}
        for cc_id, n in (DpColaborador.objects.filter(status="ativo")
                         .values_list("centro_custo_id").annotate(n=Count("id"))):
            ativos[cc_id] = n
        filhos_de = {}
        for c in todos:
            filhos_de.setdefault(c.pai_id, []).append(c)

        def montar(c):
            filhos = [montar(f) for f in sorted(filhos_de.get(c.id, []), key=lambda x: x.nome)]
            proprios = ativos.get(c.id, 0)
            return {
                "id": str(c.id), "codigo": c.codigo, "nome": c.nome,
                "nome_curto": (c.nome_curto if c.pai_id else c.nome), "ativo": c.ativo,
                "colaboradores_ativos": proprios,
                "total_com_filhos": proprios + sum(f["total_com_filhos"] for f in filhos),
                "filhos": filhos,
            }

        raizes = sorted(filhos_de.get(None, []), key=lambda x: (x.codigo, x.nome))
        return Response([montar(r) for r in raizes])

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request, "criar", "dp_centro_custo", obj.id, depois={"codigo": obj.codigo, "nome": obj.nome})

    def perform_update(self, serializer):
        antes = {"codigo": serializer.instance.codigo, "nome": serializer.instance.nome,
                 "ativo": serializer.instance.ativo}
        obj = serializer.save()
        audit(self.request, "editar", "dp_centro_custo", obj.id, antes=antes,
              depois={"codigo": obj.codigo, "nome": obj.nome, "ativo": obj.ativo})


class DpCargoViewSet(viewsets.ModelViewSet):
    queryset = DpCargo.objects.all()
    serializer_class = DpCargoSerializer
    permission_classes = _PERM

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request, "criar", "dp_cargo", obj.id,
              depois={"nome": obj.nome, "salario_base": obj.salario_base})

    def perform_update(self, serializer):
        antes = {"nome": serializer.instance.nome, "salario_base": serializer.instance.salario_base}
        obj = serializer.save()
        audit(self.request, "editar", "dp_cargo", obj.id, antes=antes,
              depois={"nome": obj.nome, "salario_base": obj.salario_base})


def _data_iso(v):
    """'2026-07-01' -> date, e qualquer coisa fora disso -> None (sem explodir)."""
    try:
        return datetime.strptime((v or "").strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def filtrar_quadro(qs, params):
    """Filtros do Quadro de Pessoal, compartilhados pela listagem e pelo EXPORT.

    Estava duplicado: a tela filtrava por busca/regime/cc/unidade e o export só
    olhava `status`, então "exportar com filtro aplicado" baixava o quadro
    inteiro. Uma função só garante que os dois não voltem a divergir.
    """
    busca = (params.get("busca") or "").strip()
    if busca:
        f = Q(nome__icontains=busca) | Q(cpf__icontains=busca)
        if busca.isdigit():
            f |= Q(matricula=int(busca))
        qs = qs.filter(f)
    if params.get("regime"):
        qs = qs.filter(regime=params["regime"])
    if params.get("status"):
        qs = qs.filter(status=params["status"])
    if params.get("cc"):
        cc = DpCentroCusto.objects.filter(pk=params["cc"]).first()
        # escolher o núcleo traz os subnúcleos junto (mesma regra da folha)
        qs = qs.filter(centro_custo_id__in=cc.descendentes_ids()) if cc else qs.none()
    if params.get("supervisor"):
        qs = qs.filter(supervisor_id=params["supervisor"])
    if params.get("coordenador"):
        qs = qs.filter(coordenador_id=params["coordenador"])
    if params.get("unidade"):
        qs = qs.filter(unidade=params["unidade"])

    # PERÍODO POR EVENTO: "quem entrou/saiu entre tais datas".
    # `evento` escolhe qual data filtrar — sem ele, um intervalo sozinho seria
    # ambíguo (data de quê?). Quem não tem a data do evento sai do resultado:
    # ativo não tem demissão, e listá-lo numa busca por desligados seria ruído.
    evento = (params.get("evento") or "").strip()
    de, ate = (params.get("de") or "").strip(), (params.get("ate") or "").strip()
    if evento in ("admissao", "desligamento") and (de or ate):
        campo = "data_admissao" if evento == "admissao" else "data_demissao"
        qs = qs.exclude(**{f"{campo}__isnull": True})
        # data malformada vira filtro ignorado, não 500 — o operador digita
        # direto na URL e no input de data enquanto ainda está incompleto
        if _data_iso(de):
            qs = qs.filter(**{f"{campo}__gte": _data_iso(de)})
        if _data_iso(ate):
            qs = qs.filter(**{f"{campo}__lte": _data_iso(ate)})
    return qs


class DpColaboradorViewSet(viewsets.ModelViewSet):
    """Quadro de pessoal. list aceita: ?busca= (nome/matrícula/cpf), ?regime=,
    ?status=, ?cc= (uuid), ?unidade=, ?limit/?offset — devolve {total, items}."""
    # prefetch dos dependentes: o serializer resume a situação do salário-família
    # em cada linha do quadro — sem isso seria uma query por colaborador
    queryset = DpColaborador.objects.select_related(
        "centro_custo", "cargo", "supervisor", "coordenador",
        # o vínculo de transferência é OneToOne nos dois sentidos: sem
        # select_related seriam 2 queries por linha do quadro
        "transferencia_saida__destino", "transferencia_entrada__origem",
    ).prefetch_related("dependentes").all()
    serializer_class = DpColaboradorSerializer
    permission_classes = _PERM

    def list(self, request, *args, **kwargs):
        qs = filtrar_quadro(
            filtrar_colaboradores(self.get_queryset(), request.user),
            request.query_params)

        # Filtro por situação do salário-família (?sf=pendente|regular|…).
        # A situação depende de idade e de datas de comprovação, que o ORM não
        # resolve num filter — então avaliamos em Python. Pré-filtramos por
        # "tem dependente ativo" no banco para não varrer o quadro inteiro.
        # Fica FORA de filtrar_quadro porque depende do serializer.
        sf = (request.query_params.get("sf") or "").strip()
        if sf:
            qs = qs.filter(dependentes__ativo=True).distinct()
            ser = self.get_serializer()
            ids = [c.id for c in qs.prefetch_related("dependentes")
                   if ser.get_salario_familia(c)["situacao"] == sf]
            qs = self.get_queryset().filter(id__in=ids)

        total = qs.count()
        try:
            limit = min(int(request.query_params.get("limit", 50)), 500)
            offset = max(int(request.query_params.get("offset", 0)), 0)
        except ValueError:
            limit, offset = 50, 0
        items = self.get_serializer(qs[offset:offset + limit], many=True).data
        return Response({"total": total, "items": items})

    def perform_create(self, serializer):
        regime = serializer.validated_data.get("regime")
        matricula = serializer.validated_data.get("matricula") or proxima_matricula(regime)
        obj = serializer.save(matricula=matricula)
        DpEvento.objects.create(
            colaborador=obj, tipo="admissao",
            data_efeito=obj.data_admissao or obj.data_entrada or date.today(),
            payload={"regime": obj.regime, "cc": obj.centro_custo.nome if obj.centro_custo_id else None},
            autor=_quem(self.request),
        )
        audit(self.request, "criar", "dp_colaborador", obj.id, depois=_snap(obj))

    def perform_update(self, serializer):
        inst = serializer.instance
        antes = _snap(inst)
        cc_antes, sal_antes = inst.centro_custo_id, inst.salario_bruto
        obj = serializer.save()
        depois = _snap(obj)
        if antes != depois:
            audit(self.request, "editar", "dp_colaborador", obj.id, antes=antes, depois=depois)
            hoje = date.today()
            if obj.centro_custo_id != cc_antes:
                DpEvento.objects.create(colaborador=obj, tipo="transferencia_cc", data_efeito=hoje,
                                        payload={"de": antes["centro_custo"], "para": depois["centro_custo"]},
                                        autor=_quem(self.request))
            if obj.salario_bruto != sal_antes:
                DpEvento.objects.create(colaborador=obj, tipo="reajuste", data_efeito=hoje,
                                        payload={"de": sal_antes, "para": obj.salario_bruto},
                                        autor=_quem(self.request))

    def destroy(self, request, *args, **kwargs):
        # Sem hard-delete no quadro: histórico é sagrado. Use /desligar.
        return Response({"detail": "Colaborador não é excluído — use o desligamento."},
                        status=status.HTTP_405_METHOD_NOT_ALLOWED)

    @action(detail=True, methods=["post"])
    def desligar(self, request, pk=None):
        """Desligamento: {data_demissao: 'YYYY-MM-DD', observacao?}. Marca inativo
        + evento de desligamento (verbas rescisórias detalhadas ficam pra F2)."""
        obj = self.get_object()
        data_str = request.data.get("data_demissao")
        try:
            data_dem = datetime.strptime(data_str, "%Y-%m-%d").date() if data_str else date.today()
        except ValueError:
            return Response({"detail": "data_demissao inválida (use YYYY-MM-DD)."},
                            status=status.HTTP_400_BAD_REQUEST)
        antes = _snap(obj)
        obj.status = "inativo"
        obj.data_demissao = data_dem
        obj.save()
        DpEvento.objects.create(colaborador=obj, tipo="desligamento", data_efeito=data_dem,
                                payload={"observacao": request.data.get("observacao", "")},
                                autor=_quem(request))
        audit(request, "desligar", "dp_colaborador", obj.id, antes=antes, depois=_snap(obj))
        return Response(self.get_serializer(obj).data)

    # ─────────────── afastamentos e suspensões ───────────────

    @action(detail=True, methods=["get", "post"])
    def afastamentos(self, request, pk=None):
        """GET lista; POST registra. Só CLT tem efeito na folha — para os demais
        regimes o registro fica como histórico, sem mexer no cálculo."""
        colab = self.get_object()
        if request.method == "GET":
            return Response([_afastamento_json(a) for a in colab.afastamentos.all()])

        tipo = request.data.get("tipo")
        if tipo not in dict(DpAfastamento.TIPOS):
            return Response({"detail": "Tipo de afastamento inválido."}, status=400)
        inicio = _data_ou_none(request.data.get("data_inicio"))
        if not inicio:
            return Response({"detail": "Informe a data de início."}, status=400)
        retorno = _data_ou_none(request.data.get("data_retorno"))
        prevista = _data_ou_none(request.data.get("data_prevista_retorno"))
        if retorno and retorno < inicio:
            return Response({"detail": "O retorno não pode ser antes do início."}, status=400)

        # estabilidade: acidente conta do RETORNO; maternidade, do início
        meses = REGRAS_AFASTAMENTO.get(tipo, {}).get("estabilidade_meses")
        estab = _data_ou_none(request.data.get("estabilidade_ate"))
        if not estab and meses:
            base = retorno or prevista or inicio
            ano_e, mes_e = base.year, base.month + meses
            while mes_e > 12:
                mes_e -= 12
                ano_e += 1
            from calendar import monthrange
            estab = date(ano_e, mes_e, min(base.day, monthrange(ano_e, mes_e)[1]))

        a = DpAfastamento.objects.create(
            colaborador=colab, tipo=tipo, data_inicio=inicio,
            data_prevista_retorno=prevista, data_retorno=retorno,
            estabilidade_ate=estab,
            observacao=(request.data.get("observacao") or "")[:250],
            registrado_por=_quem(request))
        audit(request, "afastamento", "dp_afastamento", a.id, colaborador=colab,
              depois={"colaborador": colab.nome, "tipo": a.get_tipo_display(),
                      "inicio": f"{inicio:%d/%m/%Y}",
                      "retorno": f"{retorno:%d/%m/%Y}" if retorno else "em curso",
                      "estabilidade": f"{estab:%d/%m/%Y}" if estab else "—"})
        return Response(_afastamento_json(a), status=201)

    @action(detail=True, methods=["patch", "delete"],
            url_path=r"afastamentos/(?P<af_id>[^/.]+)")
    def afastamento(self, request, pk=None, af_id=None):
        """PATCH registra o retorno (ou corrige); DELETE remove."""
        colab = self.get_object()
        a = colab.afastamentos.filter(pk=af_id).first()
        if not a:
            return Response({"detail": "Afastamento não encontrado."}, status=404)

        if request.method == "DELETE":
            audit(request, "excluir", "dp_afastamento", a.id, colaborador=colab,
                  antes={"tipo": a.get_tipo_display(), "inicio": f"{a.data_inicio:%d/%m/%Y}"},
                  depois={"colaborador": colab.nome, "situacao": "removido"})
            a.delete()
            return Response({"detail": "Afastamento removido."})

        antes = {"retorno": str(a.data_retorno or "em curso")}
        for campo in ("data_retorno", "data_prevista_retorno", "estabilidade_ate"):
            if campo in request.data:
                setattr(a, campo, _data_ou_none(request.data.get(campo)))
        if "observacao" in request.data:
            a.observacao = (request.data.get("observacao") or "")[:250]
        if a.data_retorno and a.data_retorno < a.data_inicio:
            return Response({"detail": "O retorno não pode ser antes do início."}, status=400)
        a.save()
        audit(request, "afastamento", "dp_afastamento", a.id, colaborador=colab, antes=antes,
              depois={"colaborador": colab.nome, "tipo": a.get_tipo_display(),
                      "retorno": str(a.data_retorno or "em curso")})
        return Response(_afastamento_json(a))

    # ─────────────── transferência de contrato ───────────────

    @action(detail=True, methods=["get", "post", "delete"])
    def transferencia(self, request, pk=None):
        """Registra que ESTA ficha é a continuação de outra matrícula.

        A casa numera matrícula por regime, então efetivar alguém obriga a
        abrir cadastro novo e encerrar o antigo. Sem registrar o vínculo, o
        painel lê os dois movimentos como um desligamento e uma admissão de
        pessoas diferentes, e o turnover incha com gente que nunca saiu.

        POST no DESTINO (a matrícula nova), informando a de origem.
        """
        destino = self.get_object()

        if request.method == "GET":
            return Response(self.get_serializer(destino).data.get("transferencia"))

        if request.method == "DELETE":
            t = getattr(destino, "transferencia_entrada", None)
            if not t:
                return Response({"detail": "Esta ficha não tem transferência registrada."},
                                status=404)
            origem = t.origem
            # devolve os dependentes que vieram nesta transferência — sem isso
            # o salário-família passaria a ser pago pela ficha errada
            devolvidos = 0
            if t.dependentes_ids:
                devolvidos = DpDependente.objects.filter(
                    id__in=t.dependentes_ids, colaborador=destino).update(colaborador=origem)
            t.delete()
            audit(request, "desfazer_transferencia", "dp_transferencia", destino.id,
                  colaborador=destino,
                  depois={"colaborador": destino.nome,
                          "matricula_origem": origem.matricula,
                          "matricula_destino": destino.matricula,
                          "dependentes_devolvidos": devolvidos})
            return Response({"detail": "Transferência desfeita.",
                             "dependentes_devolvidos": devolvidos})

        # ── POST ──
        origem = None
        if request.data.get("origem_id"):
            origem = DpColaborador.objects.filter(pk=request.data["origem_id"]).first()
        elif request.data.get("origem_matricula"):
            try:
                origem = DpColaborador.objects.filter(
                    matricula=int(request.data["origem_matricula"])).first()
            except (TypeError, ValueError):
                origem = None
        if not origem:
            return Response({"detail": "Informe a matrícula de origem (o contrato anterior)."},
                            status=400)
        if origem.pk == destino.pk:
            return Response({"detail": "A matrícula de origem não pode ser a mesma."}, status=400)
        if getattr(origem, "transferencia_saida", None):
            return Response({"detail": f"A matrícula {origem.matricula} já foi transferida "
                                       "para outra ficha."}, status=409)
        if getattr(destino, "transferencia_entrada", None):
            return Response({"detail": "Esta ficha já tem um contrato anterior vinculado."},
                            status=409)
        if origem.status == "ativo":
            # se as duas ficarem ativas a pessoa conta DUAS vezes no quadro
            return Response(
                {"detail": f"A matrícula {origem.matricula} ainda está ATIVA. Registre o "
                           "desligamento dela primeiro — senão a pessoa passa a contar duas "
                           "vezes no quadro de pessoal."}, status=409)

        data_efeito = (_data_ou_none(request.data.get("data_efeito"))
                       or destino.data_admissao or origem.data_demissao or date.today())
        mover = request.data.get("mover_dependentes")
        mover = True if mover is None else bool(mover)

        with transaction.atomic():
            movidos = []
            if mover:
                # não duplica: só vem quem ainda não existe na ficha nova
                ja_la = {(d.nome.strip().lower(), d.data_nascimento)
                         for d in destino.dependentes.all()}
                for d in origem.dependentes.filter(ativo=True):
                    if (d.nome.strip().lower(), d.data_nascimento) in ja_la:
                        continue
                    d.colaborador = destino
                    d.save(update_fields=["colaborador", "updated_at"])
                    movidos.append(str(d.id))

            t = DpTransferenciaContrato.objects.create(
                origem=origem, destino=destino, data_efeito=data_efeito,
                motivo=(request.data.get("motivo") or "")[:200],
                dependentes_movidos=len(movidos), dependentes_ids=movidos,
                registrado_por=_quem(request))
            audit(request, "transferencia_contrato", "dp_transferencia", t.id,
                  colaborador=destino,
                  depois={"colaborador": destino.nome,
                          "de": f"matrícula {origem.matricula} ({origem.get_regime_display()})",
                          "para": f"matrícula {destino.matricula} ({destino.get_regime_display()})",
                          "data": f"{data_efeito:%d/%m/%Y}",
                          "motivo": t.motivo or "—",
                          "dependentes_movidos": len(movidos)})

        destino.refresh_from_db()
        return Response({
            "transferencia": self.get_serializer(destino).data.get("transferencia"),
            "dependentes_movidos": len(movidos),
        }, status=201)

    # ─────────────── dependentes (salário-família) ───────────────

    @action(detail=True, methods=["get", "post"])
    def dependentes(self, request, pk=None):
        """GET lista; POST cadastra. O que decide a cota é a DATA DE NASCIMENTO
        (a cota morre no mês em que o dependente faz 14 anos) e a validade das
        comprovações — por isso nada aqui é uma contagem solta."""
        colab = self.get_object()
        if request.method == "GET":
            return Response([_dependente_json(d, colab)
                             for d in colab.dependentes.all()])

        nome = (request.data.get("nome") or "").strip()
        if not nome:
            return Response({"detail": "Informe o nome do dependente."}, status=400)
        nasc = _data_ou_none(request.data.get("data_nascimento"))
        if not nasc:
            return Response({"detail": "Informe a data de nascimento (AAAA-MM-DD). "
                                       "É ela que define até quando a cota é devida."},
                            status=400)
        if nasc > date.today():
            return Response({"detail": "Data de nascimento no futuro."}, status=400)

        dep = DpDependente.objects.create(
            colaborador=colab, nome=nome[:150], data_nascimento=nasc,
            tipo=(request.data.get("tipo") or "filho"),
            cpf=_cpf(request.data.get("cpf"))[:14],
            invalido=bool(request.data.get("invalido")),
            vacinacao_valida_ate=_data_ou_none(request.data.get("vacinacao_valida_ate")),
            frequencia_escolar_valida_ate=_data_ou_none(
                request.data.get("frequencia_escolar_valida_ate")),
            observacao=(request.data.get("observacao") or "")[:250],
        )
        audit(request, "criar", "dp_dependente", dep.id, colaborador=colab,
              depois={"colaborador": colab.nome, "dependente": dep.nome,
                      "nascimento": f"{nasc:%d/%m/%Y}", "tipo": dep.get_tipo_display()})
        return Response(_dependente_json(dep, colab), status=201)

    @action(detail=True, methods=["patch", "delete"],
            url_path=r"dependentes/(?P<dep_id>[^/.]+)")
    def dependente(self, request, pk=None, dep_id=None):
        """PATCH edita; DELETE inativa (não apaga: quem já recebeu cota precisa
        continuar existindo no histórico da folha)."""
        colab = self.get_object()
        dep = colab.dependentes.filter(pk=dep_id).first()
        if not dep:
            return Response({"detail": "Dependente não encontrado."}, status=404)

        if request.method == "DELETE":
            dep.ativo = False
            dep.save(update_fields=["ativo", "updated_at"])
            audit(request, "excluir", "dp_dependente", dep.id, colaborador=colab,
                  antes={"dependente": dep.nome, "situacao": "ativo"},
                  depois={"colaborador": colab.nome, "dependente": dep.nome,
                          "situacao": "inativado (histórico preservado)"})
            return Response(_dependente_json(dep, colab))

        antes = {"nome": dep.nome, "tipo": dep.get_tipo_display(),
                 "vacinacao": str(dep.vacinacao_valida_ate or "—"),
                 "frequencia": str(dep.frequencia_escolar_valida_ate or "—"),
                 "situacao": "ativo" if dep.ativo else "inativo"}
        if "nome" in request.data:
            dep.nome = (request.data.get("nome") or dep.nome)[:150]
        if "tipo" in request.data:
            dep.tipo = request.data.get("tipo") or dep.tipo
        if "cpf" in request.data:
            dep.cpf = _cpf(request.data.get("cpf"))[:14]
        if "invalido" in request.data:
            dep.invalido = bool(request.data.get("invalido"))
        if "ativo" in request.data:
            dep.ativo = bool(request.data.get("ativo"))
        if "observacao" in request.data:
            dep.observacao = (request.data.get("observacao") or "")[:250]
        for campo in ("data_nascimento", "vacinacao_valida_ate",
                      "frequencia_escolar_valida_ate"):
            if campo in request.data:
                valor = _data_ou_none(request.data.get(campo))
                if campo == "data_nascimento" and not valor:
                    return Response({"detail": "Data de nascimento inválida."}, status=400)
                setattr(dep, campo, valor)
        dep.save()
        audit(request, "editar", "dp_dependente", dep.id, colaborador=colab,
              antes=antes,
              depois={"colaborador": colab.nome, "nome": dep.nome,
                      "tipo": dep.get_tipo_display(),
                      "vacinacao": str(dep.vacinacao_valida_ate or "—"),
                      "frequencia": str(dep.frequencia_escolar_valida_ate or "—"),
                      "situacao": "ativo" if dep.ativo else "inativo"})
        return Response(_dependente_json(dep, colab))

    # ─────────────── documentos do colaborador (contrato em PDF) ───────────────

    @action(detail=True, methods=["get", "post"], parser_classes=[MultiPartParser, FormParser])
    def documentos(self, request, pk=None):
        """GET lista os documentos; POST envia um novo (multipart: arquivo,
        tipo?, descricao?). Só PDF, até o limite do settings."""
        colab = self.get_object()
        if request.method == "GET":
            return Response([_doc_json(d) for d in colab.documentos.all()])

        arq = request.FILES.get("arquivo")
        if not arq:
            return Response({"detail": "Anexe o arquivo PDF."}, status=400)
        if not arq.name.lower().endswith(".pdf"):
            return Response({"detail": "Só aceitamos PDF por aqui."}, status=400)
        limite = getattr(settings, "DP_UPLOAD_MAX_BYTES", 25 * 1024 * 1024)
        if arq.size > limite:
            return Response({"detail": f"Arquivo grande demais (máximo {limite // (1024*1024)} MB)."},
                            status=400)
        # confere a assinatura do PDF (extensão trocada não passa)
        cabecalho = arq.read(5)
        arq.seek(0)
        if cabecalho[:4] != b"%PDF":
            return Response({"detail": "O arquivo não parece um PDF válido."}, status=400)

        tipo = request.data.get("tipo") or "contrato"
        doc = DpDocumento.objects.create(
            colaborador=colab, tipo=tipo, arquivo=arq,
            nome_original=arq.name[:255], tamanho=arq.size,
            descricao=(request.data.get("descricao") or "")[:200],
            enviado_por=_quem(request),
        )
        audit(request, "anexar", "dp_documento", doc.id, colaborador=colab,
              depois={"colaborador": colab.nome, "tipo": doc.get_tipo_display(),
                      "arquivo": doc.nome_original, "tamanho_kb": round(doc.tamanho / 1024)})
        return Response(_doc_json(doc), status=201)

    @action(detail=True, methods=["get"], url_path=r"documentos/(?P<doc_id>[^/.]+)/baixar")
    def baixar_documento(self, request, pk=None, doc_id=None):
        """Download autenticado — o arquivo NUNCA fica exposto pelo nginx."""
        colab = self.get_object()
        doc = colab.documentos.filter(pk=doc_id).first()
        if not doc:
            return Response({"detail": "Documento não encontrado."}, status=404)
        try:
            f = doc.arquivo.open("rb")
        except (FileNotFoundError, ValueError):
            return Response({"detail": "Arquivo sumiu do disco — reenvie o documento."}, status=410)
        resp = FileResponse(f, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{doc.nome_original}"'
        resp["Access-Control-Expose-Headers"] = "Content-Disposition"
        return resp

    @action(detail=True, methods=["delete"], url_path=r"documentos/(?P<doc_id>[^/.]+)")
    def remover_documento(self, request, pk=None, doc_id=None):
        colab = self.get_object()
        doc = colab.documentos.filter(pk=doc_id).first()
        if not doc:
            return Response({"detail": "Documento não encontrado."}, status=404)
        audit(request, "excluir", "dp_documento", doc.id, colaborador=colab,
              antes={"colaborador": colab.nome, "tipo": doc.get_tipo_display(),
                     "arquivo": doc.nome_original})
        doc.arquivo.delete(save=False)
        doc.delete()
        return Response(status=204)

    @action(detail=True, methods=["get"])
    def eventos(self, request, pk=None):
        obj = self.get_object()
        return Response([
            {"tipo": e.tipo, "data_efeito": str(e.data_efeito), "payload": e.payload,
             "autor": e.autor, "created_at": e.created_at.isoformat()}
            for e in obj.eventos.all()[:100]
        ])

    @action(detail=False, methods=["get"])
    def proxima_matricula(self, request):
        regime = request.query_params.get("regime", "clt")
        return Response({"matricula": proxima_matricula(regime)})

    @action(detail=False, methods=["get"])
    def resumo(self, request):
        """KPIs do quadro pro topo da tela."""
        qs = filtrar_colaboradores(DpColaborador.objects.all(), request.user)
        ativos = qs.filter(status="ativo")
        por_regime = {r: ativos.filter(regime=r).count() for r, _ in
                      [("estagiario", ""), ("clt", ""), ("associado", ""), ("pj", "")]}
        return Response({
            "ativos": ativos.count(),
            "inativos": qs.filter(status="inativo").count(),
            "por_regime": por_regime,
        })


class DpLiderancaViewSet(viewsets.ModelViewSet):
    """Catálogo de supervisores e coordenadores. ?papel=supervisor|coordenador
    filtra por função; ?ativo=1 esconde as inativas."""
    queryset = DpLideranca.objects.select_related("centro_custo").all()
    serializer_class = DpLiderancaSerializer
    permission_classes = _PERM

    def get_queryset(self):
        qs = super().get_queryset()
        papel = self.request.query_params.get("papel")
        if papel == "supervisor":
            qs = qs.filter(e_supervisor=True)
        elif papel == "coordenador":
            qs = qs.filter(e_coordenador=True)
        if self.request.query_params.get("ativo") in ("1", "true"):
            qs = qs.filter(ativo=True)
        return qs

    def _snapshot(self, obj):
        return {"nome": obj.nome, "e_supervisor": obj.e_supervisor,
                "e_coordenador": obj.e_coordenador, "ativo": obj.ativo,
                "centro_custo": obj.centro_custo.nome if obj.centro_custo_id else None,
                "email": obj.email}

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request, "criar", "dp_lideranca", obj.id, depois=self._snapshot(obj))

    def perform_update(self, serializer):
        antes = self._snapshot(serializer.instance)
        obj = serializer.save()
        audit(self.request, "editar", "dp_lideranca", obj.id, antes=antes,
              depois=self._snapshot(obj))

    def perform_destroy(self, instance):
        # liderança com gente vinculada não some — inativa (histórico é sagrado)
        if instance.supervisionados.exists() or instance.coordenados.exists():
            instance.ativo = False
            instance.save(update_fields=["ativo"])
            audit(self.request, "editar", "dp_lideranca", instance.id,
                  antes={"nome": instance.nome, "ativo": True},
                  depois={"nome": instance.nome, "ativo": False})
            return
        audit(self.request, "excluir", "dp_lideranca", instance.id,
              antes=self._snapshot(instance))
        instance.delete()

    @action(detail=False, methods=["get"])
    def equipe(self, request):
        """Quantas pessoas cada liderança tem hoje (aba de catálogo)."""
        sup = dict(DpColaborador.objects.filter(status="ativo", supervisor__isnull=False)
                   .values_list("supervisor_id").annotate(n=Count("id")))
        coord = dict(DpColaborador.objects.filter(status="ativo", coordenador__isnull=False)
                     .values_list("coordenador_id").annotate(n=Count("id")))
        return Response({str(k): {"supervisionados": sup.get(k, 0), "coordenados": coord.get(k, 0)}
                         for k in set(sup) | set(coord)})


@api_view(["GET"])
@permission_classes(_PERM)
def dp_audit_list(request):
    """Trilha de auditoria do DP (paginada): ?limit/?offset/?entidade=
    /?usuario= (quem executou) /?colaborador= (uuid da pessoa afetada)
    /?busca= (texto livre no nome da pessoa ou no usuário)."""
    qs = DpAuditLog.objects.all()
    if request.query_params.get("entidade"):
        qs = qs.filter(entidade=request.query_params["entidade"])
    if request.query_params.get("usuario"):
        qs = qs.filter(usuario__iexact=request.query_params["usuario"])
    if request.query_params.get("colaborador"):
        qs = qs.filter(colaborador_id=request.query_params["colaborador"])
    busca = (request.query_params.get("busca") or "").strip()
    if busca:
        qs = qs.filter(Q(colaborador_nome__icontains=busca) | Q(usuario__icontains=busca))
    total = qs.count()
    try:
        limit = min(int(request.query_params.get("limit", 50)), 500)
        offset = max(int(request.query_params.get("offset", 0)), 0)
    except ValueError:
        limit, offset = 50, 0
    # devolve TRADUZIDO (linguagem humana) — a UI nao mostra JSON
    return Response({"total": total,
                     "items": [humanizar(a) for a in qs[offset:offset + limit]]})


@api_view(["GET"])
@permission_classes(_PERM)
def dp_audit_filtros(request):
    """Alimenta os filtros da auditoria: quem já mexeu no módulo e quais pessoas
    têm histórico registrado."""
    usuarios = sorted({u for u in DpAuditLog.objects.values_list("usuario", flat=True).distinct() if u})
    pessoas = (DpAuditLog.objects.filter(colaborador__isnull=False)
               .values("colaborador_id", "colaborador_nome").distinct())
    vistos, lista = set(), []
    for p in pessoas:
        cid = str(p["colaborador_id"])
        if cid in vistos:
            continue
        vistos.add(cid)
        lista.append({"id": cid, "nome": p["colaborador_nome"]})
    lista.sort(key=lambda x: x["nome"])
    entidades = sorted({e for e in DpAuditLog.objects.values_list("entidade", flat=True).distinct() if e})
    return Response({"usuarios": usuarios, "colaboradores": lista, "entidades": entidades})


# ─────────────────────────── IMPORTADOR DA PLANILHA ───────────────────────────

_REGIME_MAP = {
    "estagiário (tce)": "estagiario", "estagiario (tce)": "estagiario",
    "estagiário": "estagiario", "clt": "clt", "associado": "associado", "pj": "pj",
}


def _norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def _cpf(v) -> str:
    """CPF/CNPJ da planilha, com o zero à esquerda de volta.

    O Excel guarda documento como NÚMERO, então 063.395.424-17 chega aqui como
    63395424417 e o zero já sumiu antes do código ver o valor. Ficava um CPF de
    10 dígitos que não casa com nada — foi o que impediu 29 pessoas de bater com
    o extrato da contabilidade, e mascarou 83 registros até 13/08/2026.

    Completar à esquerda é seguro porque o comprimento é fixo: 11 para CPF, 14
    para CNPJ (PJ). Documento maior que 14 fica intocado — aí não é zero
    perdido, é dado errado, e adivinhar seria pior que deixar visível.
    """
    d = "".join(c for c in _norm(v) if c.isdigit())
    if not d:
        return ""
    if len(d) <= 11:
        return d.zfill(11)
    if len(d) <= 14:
        return d.zfill(14)
    return d


def _data(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _lideranca(nome: str, papel: str):
    """Acha (ou cria) a liderança no catálogo pelo nome vindo da planilha."""
    nome = (nome or "").strip()
    if not nome:
        return None
    lid = DpLideranca.objects.filter(nome__iexact=nome).first()
    if lid is None:
        lid = DpLideranca.objects.create(nome=nome, e_supervisor=(papel == "sup"),
                                         e_coordenador=(papel == "coord"))
    else:
        campo = "e_supervisor" if papel == "sup" else "e_coordenador"
        if not getattr(lid, campo):
            setattr(lid, campo, True)
            lid.save(update_fields=[campo])
    return lid


@api_view(["POST"])
@permission_classes(_PERM)
def dp_importar(request):
    """Importa a planilha REAL "Controle de Pessoal DP - CC.xlsx":
    CONFIG (CCs+códigos) → TB_Cargos → TB_Colaboradores (upsert por matrícula)
    → Desligados (data de demissão). Idempotente: rodar de novo atualiza.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    f = request.FILES.get("arquivo")
    if not f:
        return Response({"detail": "Envie o arquivo em 'arquivo'."}, status=400)
    try:
        wb = load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
    except Exception:
        return Response({"detail": "Arquivo inválido — envie o .xlsx do Controle de Pessoal."}, status=400)

    resumo = {"ccs": 0, "cargos": 0, "colaboradores_novos": 0,
              "colaboradores_atualizados": 0, "desligados_marcados": 0,
              # eventos de saída criados — contados à parte de "marcados",
              # porque o cadastro pode já estar certo e o evento faltar
              "desligamentos_registrados": 0, "avisos": []}

    with transaction.atomic():
        # 1) CONFIG → Centros de Custo (colunas B=CCs, C=Cód)
        if "CONFIG" in wb.sheetnames:
            ws = wb["CONFIG"]
            for row in ws.iter_rows(min_row=3, values_only=True):
                nome, cod = _norm(row[1] if len(row) > 1 else None), _norm(row[2] if len(row) > 2 else None)
                if not nome:
                    continue
                DpCentroCusto.objects.update_or_create(
                    nome=nome, defaults={"codigo": int(cod) if cod.isdigit() else 0})
                resumo["ccs"] += 1

        # 2) TB_Cargos → plano de cargos (r3=header: Área, Cargo, Salário, Dias, Carga)
        if "TB_Cargos" in wb.sheetnames:
            ws = wb["TB_Cargos"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                area, nome = _norm(row[0]), _norm(row[1] if len(row) > 1 else None)
                if not nome:
                    continue
                DpCargo.objects.update_or_create(nome=nome, defaults={
                    "area": area, "salario_base": _num(row[2] if len(row) > 2 else 0),
                    "dias_mes": int(_num(row[3] if len(row) > 3 else 30) or 30),
                    "carga_horaria_mes": int(_num(row[4] if len(row) > 4 else 220) or 220),
                })
                resumo["cargos"] += 1

        # 3) TB_Colaboradores (r3=header; colunas conforme a planilha)
        cc_por_nome = {c.nome: c for c in DpCentroCusto.objects.all()}
        cargo_por_nome = {c.nome: c for c in DpCargo.objects.all()}
        if "TB_Colaboradores" in wb.sheetnames:
            ws = wb["TB_Colaboradores"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                mat = _norm(row[1] if len(row) > 1 else None)
                nome = _norm(row[2] if len(row) > 2 else None)
                if not mat.isdigit() or not nome:
                    continue
                cc_nome = _norm(row[8] if len(row) > 8 else None)
                cc = cc_por_nome.get(cc_nome)
                if cc is None and cc_nome:
                    cc = DpCentroCusto.objects.create(nome=cc_nome, codigo=0)
                    cc_por_nome[cc_nome] = cc
                    resumo["avisos"].append(f"CC criado fora do CONFIG: {cc_nome}")
                if cc is None:
                    resumo["avisos"].append(f"Colaborador {mat} sem CC — pulado")
                    continue
                cargo_nome = _norm(row[11] if len(row) > 11 else None)
                cargo = cargo_por_nome.get(cargo_nome)
                if cargo is None and cargo_nome:
                    # Cargo que existe no quadro mas não na TB_Cargos. Antes isso
                    # gravava cargo=None EM SILÊNCIO — e como o cargo é quem
                    # define dias_mes/carga_horária, a folha dessa pessoa passava
                    # a calcular faltas pelo default (30/220) sem ninguém saber.
                    # Cria com o mesmo critério já usado pro centro de custo e
                    # avisa, pra o DP completar a TB_Cargos depois.
                    cargo = DpCargo.objects.create(
                        nome=cargo_nome, area=_norm(row[6] if len(row) > 6 else ""),
                        salario_base=0, dias_mes=30, carga_horaria_mes=220)
                    cargo_por_nome[cargo_nome] = cargo
                    resumo["avisos"].append(
                        f"Cargo criado fora da TB_Cargos: {cargo_nome} "
                        f"(sem salário base — confira dias/carga com o DP)")
                regime = _REGIME_MAP.get(_norm(row[13] if len(row) > 13 else "").lower(), "clt")
                campos = {
                    "nome": nome, "sexo": _norm(row[3]), "cpf": _cpf(row[4]),
                    "unidade": _norm(row[5]), "area": _norm(row[6]),
                    "centro_custo": cc,
                    # a planilha traz "Fulano - SUP" / "Fulano - COOR" na mesma
                    # coluna — cada nome vira/acha uma linha no catálogo
                    "supervisor": _lideranca(_norm(row[9]), "sup"),
                    "coordenador": (_lideranca(_norm(row[9]), "coord")
                                    if "COOR" in _norm(row[9]).upper() else None),
                    "equipe": _norm(row[10]),
                    "cargo": cargo, "regime": regime,
                    "status": "ativo" if _norm(row[12]).lower() == "ativo" else "inativo",
                    "data_entrada": _data(row[14] if len(row) > 14 else None),
                    "data_admissao": _data(row[15] if len(row) > 15 else None),
                    "data_demissao": _data(row[16] if len(row) > 16 else None),
                    "salario_bruto": _num(row[17] if len(row) > 17 else 0),
                    "saldo_livre": _num(row[18] if len(row) > 18 else 0),
                    "vt": _num(row[19] if len(row) > 19 else 0),
                    "opta_vt": _norm(row[20] if len(row) > 20 else "").lower() not in ("não", "nao", "n"),
                    "va": _num(row[21] if len(row) > 21 else 0),
                    "conta_bb": _norm(row[22] if len(row) > 22 else ""),
                    "pix": _norm(row[23] if len(row) > 23 else ""),
                    "conta_caixa": _norm(row[24] if len(row) > 24 else ""),
                }
                obj, criado = DpColaborador.objects.update_or_create(
                    matricula=int(mat), defaults=campos)
                resumo["colaboradores_novos" if criado else "colaboradores_atualizados"] += 1
                if criado:
                    DpEvento.objects.create(
                        colaborador=obj, tipo="importacao",
                        data_efeito=obj.data_admissao or date.today(),
                        payload={"origem": "planilha"}, autor=_quem(request))

        # 4) Desligados → garante data de demissão/evento (r2=header)
        if "Desligados" in wb.sheetnames:
            ws = wb["Desligados"]
            for row in ws.iter_rows(min_row=3, values_only=True):
                data_dem = _data(row[0] if row else None)
                mat = _norm(row[1] if len(row) > 1 else None)
                if not mat.isdigit() or not data_dem:
                    continue
                obj = DpColaborador.objects.filter(matricula=int(mat)).first()
                if not obj:
                    continue
                # 1) o CADASTRO: só grava se estiver diferente
                if obj.status != "inativo" or obj.data_demissao != data_dem:
                    obj.status = "inativo"
                    obj.data_demissao = data_dem
                    obj.save()
                    resumo["desligados_marcados"] += 1
                # 2) o EVENTO: garantido SEMPRE, fora do if acima.
                # Antes ele vivia dentro daquele bloco e por isso quase nunca
                # nascia: a aba TB_Colaboradores já traz status "Inativo" e a
                # data de demissão, então quando a aba Desligados era lida o
                # cadastro JÁ estava correto, o if dava falso e o evento não
                # era criado. Resultado: o desligamento não existia na trilha e
                # a evolução mensal mostrava admissões sem nenhuma saída.
                if not obj.eventos.filter(tipo="desligamento").exists():
                    DpEvento.objects.create(colaborador=obj, tipo="desligamento",
                                            data_efeito=data_dem,
                                            payload={"origem": "planilha"}, autor=_quem(request))
                    resumo["desligamentos_registrados"] += 1

    wb.close()
    audit(request, "importar", "dp_importacao", "", depois=resumo)
    return Response(resumo)
